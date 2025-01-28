from django.shortcuts import render, redirect
from openpyxl import load_workbook
from gestionvencimientos.models import Actividad, Actividad_epm, Encargado

def subir_masivo_actividad_contrato(request):
    try:
        if request.method == 'POST':
            file = request.FILES['file']
            process_excel_actividad_contrato(file)
            return redirect('index_admin')
    except Exception as e:
        print(e)
    return render(request, 'carga_masiva/subir_masivo_actividad_contrato.html') 


def process_excel_actividad_contrato(file):
    try:        
        wb = load_workbook(file)
        ws = wb[wb.sheetnames[0]]
        row_count = 0
        
        for row in ws.iter_rows():
            if row_count == 0:
                row_count += 1
                continue
            actividad_contrato = Actividad()
            actividad_contrato.id = row[0].value
            actividad_contrato.nombre = row[1].value
            actividad_contrato.dias_urbano = row[2].value
            actividad_contrato.dias_rural = row[3].value
            actividad_contrato.encargado = Encargado.objects.get(id=row[4].value)
            actividad_contrato.save()

            row_count += 1
            
    except Exception as e:
        print(e)


def subir_masivo_actividad_epm(request):
    try:
        if request.method == 'POST':
            file = request.FILES['file']
            process_excel_actividad_epm(file)
            return redirect('index_admin')
    except Exception as e:
        print(e)
    return render(request, 'carga_masiva/subir_masivo_actividad_epm.html') 


def process_excel_actividad_epm(file):
    try:        
        wb = load_workbook(file)
        ws = wb[wb.sheetnames[0]]
        row_count = 0
        
        for row in ws.iter_rows():
            if row_count == 0:
                row_count += 1
                continue
            actividad_contrato = Actividad_epm()
            actividad_contrato.id = row[0].value
            actividad_contrato.nombre = row[1].value
            actividad_contrato.dias_urbano = row[2].value
            actividad_contrato.dias_rural = row[3].value
            actividad_contrato.save()
            
            row_count += 1
            
    except Exception as e:
        print(e)


def subir_masivo_encargados(request):
    try:
        if request.method == 'POST':
            file = request.FILES['file']
            process_excel_encargados(file)
            return redirect('index_admin')
    except Exception as e:
        print(e)
    return render(request, 'carga_masiva/subir_masivo_encargados.html') 


def process_excel_encargados(file):
    try:        
        wb = load_workbook(file)
        ws = wb[wb.sheetnames[0]]
        row_count = 0
        
        for row in ws.iter_rows():
            if row_count == 0:
                row_count += 1
                continue
            encargado = Encargado()
            encargado.id = row[0].value
            encargado.nombre = row[1].value
            encargado.save()
            
            row_count += 1
            
    except Exception as e:
        print(e)

