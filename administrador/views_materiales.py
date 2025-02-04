from django.http import JsonResponse
from django.shortcuts import render, redirect
from openpyxl import load_workbook
from analisis_acta.models import Materiales
from administrador.query.actividades.actividades_contrato import crear_nuevo_material, eliminar_material, actualizar_material

def materiales_permitidos_list(request):
    context={
        'materiales_permitidos': Materiales.objects.all()
    }
    return render(request, "materiales_permitidos_list.html", context)

def nuevo_material(request):
    if request.method == 'POST':
        return crear_nuevo_material(request)  # Devolver el JsonResponse directamente    
    return render(request, "nuevo_material.html")

    
def subir_masivo_materiales_contrato(request):
    try:
        if request.method == 'POST':
            file = request.FILES['file']
            process_excel_materiales_contrato(file)
            return redirect('index_admin')
    except Exception as e:
        print(e)
    return render(request, 'carga_masiva/subir_masivo_materiales_contrato.html') 


def process_excel_materiales_contrato(file):
    try:        
        wb = load_workbook(file)
        ws = wb[wb.sheetnames[0]]
        row_count = 0
        
        for row in ws.iter_rows():
            if row_count == 0:
                row_count += 1
                continue
            materiale = Materiales()
            materiale.id = row[0].value
            materiale.material = row[1].value
            materiale.save()

            row_count += 1
            
    except Exception as e:
        print(e)

def eliminar_material_contrato(request, id):
    eliminar_material(id)
    return redirect('index_admin')

def editar_material_id(request, material_id):
    if request.method == 'POST':
        try:
            return actualizar_material(request, material_id)
        except Materiales.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Material no encontrado'})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

