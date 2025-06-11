from asyncio.windows_events import NULL
from email.policy import HTTP
from django.shortcuts import redirect, render
from django.urls import reverse
from analisis_acta.models import Acta, CantidadItem, Materiales, Novedad_acta, VariableAnalisis
from django.db.models import Sum, Q, Count
from django.http import JsonResponse, HttpResponseRedirect
from openpyxl import load_workbook
from io import BytesIO

def subir_acta_revision(request):
    try:
        if request.method == 'POST':
            # Verificar si el archivo fue subido correctamente
            if 'file' not in request.FILES:
                return JsonResponse({'status': 'error', 'message': 'No file provided'}, status=400)
            file = request.FILES['file']
            # Procesar el archivo en segundo plano (en este ejemplo, se hace sincrónicamente)
            process_excel_acta(file)
            return HttpResponseRedirect(reverse('home'))  # Usamos HttpResponseRedirect en lugar de redirect
        
        # Si no es un POST, solo renderizamos la página
        return render(request, 'subir_acta_revision.html')
    
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

def process_excel_acta(file):
    try:         
        # Validar que el archivo es un Excel válido
        if not file.name.endswith('.xlsx'):
            raise ValueError("El archivo debe tener formato .xlsx")

        # Cargar el archivo en memoria usando BytesIO
        content = file.read()
        wb = load_workbook(filename=BytesIO(content))
        ws = wb.active  # Usamos la hoja activa por defecto
        row_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):  # Comenzamos desde la fila 2
            row_count += 1

            # Crear una nueva instancia del modelo Acta
            acta = Acta()

            # Asignar los valores de las filas del Excel a los campos del modelo Acta
            acta.pedido = row[0]
            acta.area_operativa = row[1]
            acta.subz = row[2]
            acta.ruta = row[3]
            acta.municipio = row[4]
            acta.contrato = row[5]
            acta.acta = row[6]
            acta.actividad = row[7]
            acta.fecha_estado = row[8]
            acta.pagina = row[9]
            acta.urbrur = row[10]
            acta.tipre = row[11]
            acta.red_interna = row[12]
            acta.tipo_operacion = row[13]
            acta.descent = row[14]
            acta.tipo = row[15]
            acta.cobro = row[16]
            acta.suminis = row[17] if row[17] is not None else '0'
            acta.item_cont = row[18] if row[18] is not None else row[17]
            acta.item_res = row[19]
            acta.cantidad = row[20] if row[20] is not None else '0'
            acta.vlr_cliente = row[21]
            acta.valor_costo = row[22]
            acta.tipo_item = row[23]

            try:
                acta.save()
            except Exception as e:
                print(f"Error al guardar la fila {row_count}: {e}")

    except Exception as e:
        print(f"Error al procesar el archivo: {e}")

def calculo_novedades_acta(request):
    pedidos_LEGA = Acta.objects.filter(actividad__in=["ALEGA", "ALECA", "ACAMN", "ALEGN"]).exclude(suminis__exact='0').exclude(suminis__exact='CALE1F')

    for pedido in pedidos_LEGA:
        if pedido.suminis not in ["215887A", "219404A"]:
            crear_novedad(pedido, "Legalización: Suministro " + pedido.suminis)
            
    pedidos_aejdo = Acta.objects.filter(actividad__in=["AEJDO"]).exclude(suminis__exact='0').exclude(
    suminis__exact='CALE1F').exclude(suminis__isnull=True)
    
    for pedido in pedidos_aejdo: 

        if pedido.suminis.endswith("P"):
            crear_novedad(pedido, "AEJDO: con suministro " + pedido.suminis)

 
    pedidos = Acta.objects.all()

    for pedido in pedidos:

        material = Materiales.objects.filter(material=pedido.suminis).exists()
        
        if pedido.suminis != '0' and material == False and pedido.suminis[0]!="C" and pedido.suminis[0]!="R":
            novedad = "Material no permitido " + str(pedido.suminis)
            crear_novedad(pedido, novedad)
        else:   
            if pedido.item_cont is None:
                continue

            cod = pedido.item_cont
            print(cod)
                
            primera_letra = cod[0]
            if primera_letra == 'A':

                if pedido.item_cont == 'A 05':
                    if int(pedido.cantidad) > 1:
                        crear_novedad(
                            pedido, 'A 05 mayor a 1')
                    
                    try:
                        # ya en la nueva logica
                        BREAKER_210954 = Acta.objects.filter(pedido=pedido).filter(
                            item_cont='210954').aggregate(suma=Sum('cantidad'))
                        
                        if BREAKER_210954['suma'] == None: #YA EN NUEVA LOGICA
                            crear_novedad(
                                pedido, 'A 05 sin 210954.')
                            
                        if BREAKER_210954['suma'] < 1: #YA EN NUEVA LOGICA
                            crear_novedad(
                                pedido, 'A 05 con 210954 menor a 1')
                    except Exception as e:
                        print(e)
                    
                    if pedido.item_cont == '210954': #YA EN NUEVA LOGICA

                        A05 = Acta.objects.filter(pedido=pedido).filter(
                            item_cont='A 05').aggregate(suma=Sum('cantidad'))
                        if BREAKER_210954['suma'] < 1:
                            crear_novedad(
                                pedido, '210954 sin A 05')

                if pedido.item_cont == 'A 04':
                    if int(pedido.cantidad) == 2:
                        BREAKER_210949 = Acta.objects.filter(pedido=pedido).filter(
                            item_cont='210949').aggregate(suma=Sum('cantidad'))
                        if BREAKER_210949['suma'] != 2:
                            crear_novedad(
                                pedido, 'A 04: 2 - 210949: ' + str(BREAKER_210949['suma']))
                    if int(pedido.cantidad) > 2:
                        nov = "Actividad: " + \
                            str(pedido.item_cont) + \
                            " con cantidad= " + str(pedido.cantidad)
                        crear_novedad(pedido, nov)
                elif int(pedido.cantidad) > 1 and pedido.item_cont != 'A 46':
                    nov = "Actividad: " + \
                        str(pedido.item_cont) + \
                        " con cantidad= " + str(pedido.cantidad)
                    crear_novedad(pedido, nov)
            
            if pedido.item_cont == 'A 06':
                # Validar que el pedido tenga un item_cont 'A 10' o 'A 11'
                validar_pedido_A06(pedido)
            

            if primera_letra == 'C' or primera_letra == 'D' or primera_letra == 'R':
                if int(pedido.cantidad) > 1:
                    nov = "Actividad: " + \
                        str(pedido.item_cont) + \
                        " con cantidad= " + str(pedido.cantidad)
                    crear_novedad(pedido, nov)

            if primera_letra == 'B':
                if pedido.item_cont == 'B 03':

                    if float(pedido.cantidad) > 60:
                        nov = "Actividad: " + \
                            str(pedido.item_cont) + \
                            " con cantidad= " + str(pedido.cantidad)
                        crear_novedad(pedido, nov)

                if pedido.item_cont == 'B 04':
                    if float(pedido.cantidad) > 13:
                        nov = "Actividad: " + \
                            str(pedido.item_cont) + \
                            " con cantidad= " + str(pedido.cantidad)
                        crear_novedad(pedido, nov)

                if pedido.item_cont == 'B 06' or pedido.item_cont == 'B 07' or pedido.item_cont == 'B 08':
                    if float(pedido.cantidad) > 5:
                        nov = "Actividad: " + \
                            str(pedido.item_cont) + \
                            " con cantidad= " + str(pedido.cantidad)
                        crear_novedad(pedido, nov)

            if pedido.item_cont == '0':
                pedido.item_cont = pedido.suminis
                pedido.save()

            pagina = pedido.pagina

            if (pagina[6:9] != '100' and pedido.urbrur == 'U') or (pagina[6:9] != '200' and pedido.urbrur == 'R'):
                try:
                    busquedad_tipo_pagina = Novedad_acta.objects.filter(
                        pedido=pedido.pedido).filter(novedad='Tipo página').count()
                    if busquedad_tipo_pagina == 0:
                        novedad = "Tipo página"
                        crear_novedad(pedido, novedad)
                except:
                    pass

            if pedido.item_cont[0:4] == 'CALE' and int(pedido.cantidad) > 1:
                crear_novedad(pedido, '')

            if pedido.actividad == 'AEJDO':

                if pedido.item_cont == "A 06":
                    if Acta.objects.filter(pedido=pedido.pedido).filter(
                            item_cont='A 39').aggregate(suma=Sum('cantidad'))['suma']==None and Acta.objects.filter(pedido=pedido.pedido).filter(
                            item_cont='A 42').aggregate(suma=Sum('cantidad'))['suma']==None:
                        crear_novedad(
                                    pedido, pedido.item_cont + ' No cobró A 39 ni A 42')                       

                if pedido.item_cont == "200410" or pedido.item_cont == "200411":
                    if int(pedido.cantidad) <= 5:

                        cant_200411 = Acta.objects.filter(pedido=pedido.pedido).filter(
                            item_cont='200411').aggregate(suma=Sum('cantidad'))['suma']
                        cant_200410 = Acta.objects.filter(pedido=pedido.pedido).filter(
                            item_cont='200410').aggregate(suma=Sum('cantidad'))['suma']

                        if cant_200411 == None:
                            cant_200411 = 0
                        if cant_200410 == None:
                            cant_200410 = 0
                        
                        if cant_200410>0 or cant_200411>0:
                            if Acta.objects.filter(pedido=pedido.pedido).filter(
                            item_cont='A 01').aggregate(suma=Sum('cantidad'))['suma']==None:
                                
                                if Acta.objects.filter(pedido=pedido.pedido).filter(
                                item_cont='A 06').aggregate(suma=Sum('cantidad'))['suma']==None:
                                    crear_novedad(
                                    pedido, pedido.item_cont + ' No cobró A 06')
                            
                            
                            

                        if (cant_200411 <= 5 and cant_200411 > 0) or (cant_200410 <= 5 and cant_200410 > 0):
                            crear_novedad(
                                pedido, '200410= ' + str(cant_200410) + ' 200411= ' + str(cant_200411))

                if pedido.item_cont == 'A 01':
                    try:
                        busquedad_a04 = Acta.objects.filter(
                            pedido=pedido.pedido).filter(item_cont='A 04').count()
                        busquedad_a05 = Acta.objects.filter(
                            pedido=pedido.pedido).filter(item_cont='A 05').count()

                        if busquedad_a04 == 0 and busquedad_a05 == 0:
                            novedad = "A 01=1, Sin A 04 y A 05"
                            crear_novedad(pedido, novedad)

                        if busquedad_a04 > 0 and busquedad_a05 > 0:
                            novedad = "A 04 y A 05 incompatibles"
                            crear_novedad(pedido, novedad)

                    except:
                        pass

                    try:
                        busquedad_A23 = Acta.objects.filter(
                            pedido=pedido.pedido).filter(item_cont='A 23').count()
                        if busquedad_A23 == 0:
                            novedad = "A 01=1, A 23=0"
                            crear_novedad(pedido, novedad)
                    except:
                        pass

                    nov = "A 01=1,"
                    busqueda_item(pedido, '200410', '200411', nov)
                    busqueda_item(pedido, '211829', 0, nov)
                    busqueda_item(pedido, '200092', '200093', nov)
                    busqueda_item(pedido, '200316', 0, nov)
                    busqueda_item(pedido, '211357', 0, nov)
                    busqueda_item(pedido, '213333', 0, nov)
                    busqueda_item(pedido, '219404', 0, nov)
                    calculo_incompatible_A01(pedido, nov)

            if pedido.tipre == 'ENESUB' and (pedido.item_cont == 'A 27' or pedido.item_cont == 'A 03' or pedido.item_cont == 'A 28' or pedido.item_cont == 'A 29'):
                nov = 'ENESUB con item ' + str(pedido.item_cont)
                crear_novedad(pedido, nov)

            if pedido.tipre == 'ENEPRE' and pedido.item_cont == 'A 44':

                nov = 'ENEPRE con item ' + str(pedido.item_cont)
                crear_novedad(pedido, nov)

            if pedido.actividad == 'DSPRE':
                if pedido.tipre == 'ENEPRE':
                    calculo_enepre(pedido)

                if pedido.item_cont == '200099':
                    if Acta.objects.filter(pedido=pedido.pedido).filter(
                                item_cont='200319').aggregate(suma=Sum('cantidad'))['suma']==None:
                            crear_novedad(
                                        pedido, pedido.item_cont + ' No cobró 200319')                                         

                if pedido.tipre == 'ENESUB':
                    calculo_enepre(pedido)

            # Obtenemos la suma para item_cont='200099'
            suma_cantidad = Acta.objects.filter(pedido=pedido.pedido, item_cont='200099').aggregate(suma=Sum('cantidad'))['suma']

            if suma_cantidad is not None and suma_cantidad > 1:
                # Antes de crear la novedad, verificamos si ya existe
                novedad_texto = 'Cobro de 200099 mayor a 1.'
                existe_novedad = Novedad_acta.objects.filter(pedido=pedido.pedido, novedad=novedad_texto).exists()

                if not existe_novedad:
                    crear_novedad(pedido, novedad_texto)
                            
            if pedido.tipre == 'ENESUB' and (pedido.item_cont == 'A 27' or pedido.item_cont == 'A 03' or pedido.item_cont == 'A 28' or pedido.item_cont == 'A 29'):
                print("error enesub")

            if pedido.item_cont == 'A 02':
                nov = "A 02=1,"
                busqueda_item(pedido, '211829', 0, nov)

            if pedido.item_cont == 'A 10' or pedido.item_cont == 'A 11':
                busqueda_item(pedido, '200410', 0, pedido.item_cont)
                busqueda_item(pedido, '211829', 0, pedido.item_cont)

            if pedido.item_cont == '211829':
                insumo = "211829"
                busqueda_insumo(pedido, insumo)

                if int(pedido.cantidad) > 1:
                    crear_novedad(pedido, str(pedido.item_cont) +
                                  ". Cantidad mayor a 1.")

            if pedido.item_cont == 'A 04':
                nov = "A 04=1,"
                busqueda_item(pedido, '210948', '210949', nov)

            if pedido.item_cont == '210948' or pedido.item_cont == '210949':

                if int(pedido.cantidad) > 2:
                    crear_novedad(pedido, str(pedido.item_cont) +
                                  ". Cantidad mayor a 2.")

                insumo = str(pedido.item_cont)
                busqueda_insumo(pedido, insumo)

            if pedido.item_cont == '210947':
                if int(pedido.cantidad) > 3:
                    crear_novedad(pedido, str(pedido.item_cont) +
                                  ". Cantidad mayor a 3.")

            if pedido.item_cont == '220683':
                if int(pedido.cantidad) > 1:
                    crear_novedad(pedido, str(pedido.item_cont) +
                                  ". Cantidad mayor a 1.")

            if pedido.item_cont == 'A 06':
                nov = "A 06=1,"
                busqueda_item(pedido, '200410', '200411', nov)
                busqueda_item(pedido, 'A 03', 'A 27', nov)


            if pedido.item_cont == '200410' or pedido.item_cont == '200411':
                insumo = str(pedido.item_cont)
                busqueda_insumo(pedido, insumo)

            if pedido.item_cont == 'A 23':
                nov = "A 23=1,"
                busqueda_item(pedido, '211673', '210947', nov)
                busqueda_item(pedido, '335931', '210947', nov)

            if pedido.item_cont == "A 23":
                    if Acta.objects.filter(pedido=pedido.pedido).filter(
                            item_cont='211673').aggregate(suma=Sum('cantidad'))['suma']==None and Acta.objects.filter(pedido=pedido.pedido).filter(
                            item_cont='335931').aggregate(suma=Sum('cantidad'))['suma']==None:
                        crear_novedad(
                                    pedido, pedido.item_cont + ' No cobró tablero')  

            if pedido.item_cont == '211673':
                insumo = str(pedido.item_cont)
                busqueda_insumo(pedido, insumo)

                if int(pedido.cantidad) > 1:
                    crear_novedad(pedido, str(pedido.item_cont) +
                                  ". Cantidad mayor a 1.")
                    
            if pedido.item_cont == '335931':
                insumo = str(pedido.item_cont)
                busqueda_insumo(pedido, insumo)

                if int(pedido.cantidad) > 1:
                    crear_novedad(pedido, str(pedido.item_cont) +
                                  ". Cantidad mayor a 1.")

            if pedido.item_cont == 'A 24':
                nov = "A 24=1,"
                busqueda_item(pedido, '211357', 0, nov)

            if pedido.item_cont == '211357':
                insumo = str(pedido.item_cont)
                busqueda_insumo(pedido, insumo)

                if int(pedido.cantidad) > 1:
                    crear_novedad(pedido, str(pedido.item_cont) +
                                  ". Cantidad mayor a 1.")

            if pedido.item_cont == 'A 25':
                nov = "A 25=1,"
                busqueda_item(pedido, '213333', 0, nov)

            if pedido.item_cont == '213333':
                insumo = str(pedido.item_cont)
                busqueda_insumo(pedido, insumo)

                if int(pedido.cantidad) > 1:
                    crear_novedad(pedido, str(pedido.item_cont) +
                                  ". Cantidad mayor a 1.")

            if pedido.item_cont == 'A 34':
                nov = "A 34=1,"
                busqueda_item(pedido, '211686', 0, nov)

            if pedido.item_cont == '211686':
                insumo = str(pedido.item_cont)
                busqueda_insumo_por_item(pedido, insumo, 'A 34')

            if pedido.item_cont == 'A 39' and pedido.actividad != "AEJDO":
                nov = "A 39=1,"
                busqueda_item(pedido, '200410', '200411', nov)
                busquedad_200410 = Acta.objects.filter(
                    pedido=pedido.pedido).filter(item_cont='200410')
                busquedad_200411 = Acta.objects.filter(
                    pedido=pedido.pedido).filter(item_cont='200411')

                for b in busquedad_200410:
                    if len(busquedad_200410) > 0:
                        if float(b.cantidad) < 8:
                            crear_novedad(
                                pedido, 'Cantidad de cable 200410 ó 200411 menor a 8')

                for b in busquedad_200411:
                    if len(busquedad_200411) > 0:
                        if float(b.cantidad) < 8:
                            crear_novedad(
                                pedido, 'Cantidad de cable 200410 ó 200411 menor a 8')

            if pedido.item_cont == 'A 42':
                nov = "A 42=1,"
                busqueda_item(pedido, '200410', '200411', nov)

            if pedido.item_cont == '200092' or pedido.item_cont == '200093' or pedido.item_cont == '200098':
                insumo = str(pedido.item_cont)
                busqueda_insumo(pedido, insumo)

            if pedido.cantidad == '0' or pedido.cantidad == "":
                crear_novedad(pedido, str(pedido.item_cont) +
                              ". Cantidad igual a cero.")

            if pedido.suminis == '200092' or pedido.suminis == '200098':
                verificar_cable_acta(pedido.pedido, '200411',
                                     '200410', pedido.suminis)

                if int(pedido.cantidad) > 1:
                    crear_novedad(pedido, str(pedido.item_cont) +
                                  ". Cantidad mayor a 1.")

            if pedido.suminis == '200092':
                comprobar_cobro_calibracion(pedido)

            if pedido.suminis == '200093':
                comprobar_cobro_calibracion(pedido)
                verificar_cable_acta(
                    pedido, '200410', '200411', pedido.suminis)

                if int(pedido.cantidad) > 1:
                    crear_novedad(pedido, str(pedido.item_cont) +
                                  ". Cantidad mayor a 1.")

            if pedido.item_cont == 'A 40':
                calculo_otros_incompatibles(pedido, 'A 41', pedido.item_cont)

            if pedido.item_cont == 'A 41':
                calculo_otros_incompatibles(pedido, 'A 40', pedido.item_cont)

            if pedido.item_cont == 'A 10' or pedido.item_cont == 'A 11':
                if int(pedido.cantidad) > 1:
                    crear_novedad(pedido, str(pedido.item_cont) +
                                  ". Cantidad mayor a 1.")
                if pedido.item_cont == 'A 10':
                    calculo_otros_incompatibles(
                        pedido, 'A 11', pedido.item_cont)
                if pedido.item_cont == 'A 11':
                    calculo_otros_incompatibles(
                        pedido, 'A 10', pedido.item_cont)  
            
                          
    
    pagina_legalizaciones()

    verificar_A03_MEDIDOR()

    verificar_y_crear_novedades_duplicadas()
    
    novedades = Novedad_acta.objects.all()

    return render(request, "analisis.html", {'novedades': novedades})

def validar_pedido_A06(pedido):
    # Buscar todos los registros asociados a este pedido
    items_pedido = Acta.objects.filter(pedido=pedido.pedido)
    
    # Verificar si existe un registro con item_cont = 'A 10' o 'A 11'
    tiene_A10_o_A11 = items_pedido.filter(item_cont__in=['A 10', 'A 11']).exists()

    # Validar que si tiene 'A 06' también debe tener 'A 10' o 'A 11'
    if not tiene_A10_o_A11:
        # Crear la novedad si no cumple la condición
        crear_novedad(pedido, "A 06 no tiene asociado A 10 o A 11")

def verificar_y_crear_novedades_duplicadas():
    # Obtener todos los registros con duplicados
    registros_duplicados = Acta.objects.values('pedido', 'item_cont').annotate(count=Count('id')).filter(count__gt=1)

    # Crear un conjunto para realizar un seguimiento de las novedades ya creadas
    novedades_creadas = set()
    cont=0

    for duplicado in registros_duplicados:
        pedido = Acta.objects.filter(pedido=duplicado['pedido']).first()
        item_cont = duplicado['item_cont']

        novedad = str(str(duplicado['pedido']) + str(item_cont) +  ' Linea duplicada')
        
        # Verificar si la novedad ya ha sido creada
        if novedad not in novedades_creadas:
            crear_novedad(pedido, str( str(item_cont) +  ' Linea duplicada'))
            novedades_creadas.add(novedad)

def pagina_legalizaciones():

    # Obtener registros con item_cont igual a C01
    registros_C01 = Acta.objects.filter(item_cont='C 01')

    # Verificar las condiciones para cada registro
    for registro in registros_C01:
        primeros_14_caracteres = registro.pagina[:14]
        cantidad_apariciones = Acta.objects.filter(
            item_cont='C 01', pagina__startswith=primeros_14_caracteres).count()

        if cantidad_apariciones == 1:
            pass
        else:
            if Novedad_acta.objects.filter(pedido=registro.pedido, novedad="Novedad C 01").count() > 0:
                continue
            crear_novedad(registro, 'Novedad C 01')

    # Obtener registros con item_cont igual a C02
    registros_C02 = Acta.objects.filter(item_cont='C 02')

    # Verificar las condiciones para cada registro
    for registro in registros_C02:
        primeros_14_caracteres = registro.pagina[:14]
        cantidad_apariciones = Acta.objects.filter(
            item_cont='C 02', pagina__startswith=primeros_14_caracteres).count()

        if 2 <= cantidad_apariciones <= 4:
            pass
        else:
            if Novedad_acta.objects.filter(pedido=registro.pedido, novedad="Novedad C 02").count() > 0:
                continue
            crear_novedad(registro, 'Novedad C 02')

    # Obtener registros con item_cont igual a C 03
    registros_C03 = Acta.objects.filter(item_cont='C 03')

    # Verificar las condiciones para cada registro
    for registro in registros_C03:
        primeros_14_caracteres = registro.pagina[:14]
        cantidad_apariciones = Acta.objects.filter(
            item_cont='C 03', pagina__startswith=primeros_14_caracteres).count()

        if 5 <= cantidad_apariciones <= 24:
            pass
        else:
            if Novedad_acta.objects.filter(pedido=registro.pedido, novedad="Novedad C 03").count() > 0:
                continue
            crear_novedad(registro, 'Novedad C 03')

    # Obtener registros con item_cont igual a C 03
    registros_C04 = Acta.objects.filter(item_cont='C 04')

    # Verificar las condiciones para cada registro
    for registro in registros_C04:
        primeros_14_caracteres = registro.pagina[:14]
        cantidad_apariciones = Acta.objects.filter(
            item_cont='C 04', pagina__startswith=primeros_14_caracteres).count()

        if cantidad_apariciones > 24:
            pass
        else:
            if Novedad_acta.objects.filter(pedido=registro.pedido, novedad="Novedad C 04").count() > 0:
                continue
            crear_novedad(registro, 'Novedad C 04')

def comprobar_cobro_calibracion(pedido):
    calibracion = Acta.objects.filter(pedido=pedido.pedido).filter(
        item_cont='CALE1F').aggregate(suma=Sum('cantidad'))['suma']
    if calibracion != 1:
        if calibracion == None:
            calibracion = 0
        crear_novedad(pedido, 'Calibración con cantidad= ' + str(calibracion))

def busqueda_insumo_por_item(pedido, insumo, item):
    try:
        pedidos = Acta.objects.filter(pedido=pedido)

        encontreinsumo = 0
        for p in pedidos:
            letra = p.item_cont[0]
            if letra == 'A':

                if p.item_cont == item:
                    encontreinsumo = 1

        if encontreinsumo == 0:
            nov = insumo + ', insumo sin actividad'
            crear_novedad(pedido, nov)

    except:
        pass

def busqueda_insumo(pedido, insumo):
    try:
        pedidos = Acta.objects.filter(pedido=pedido)

        if insumo == '200092' or insumo == '200093' or insumo == '200098':
            encontreinsumo = 0
            for p in pedidos:
                letra = p.item_cont[0]
                if letra == 'A':

                    if p.item_cont == 'A 03' or p.item_cont == 'A 44' or p.item_cont == 'A 01' or p.item_cont == 'A 27':
                        encontreinsumo = 1

            if encontreinsumo == 0:
                nov = insumo + ', insumo sin actividad'
                crear_novedad(pedido, nov)

        if insumo == '211357':
            encontreinsumo = 0
            for p in pedidos:
                letra = p.item_cont[0]
                if letra == 'A':

                    if p.item_cont == 'A 24' or p.item_cont == 'A 01':
                        encontreinsumo = 1

            if encontreinsumo == 0:
                nov = insumo + ', insumo sin actividad'
                crear_novedad(pedido, nov)

        if insumo == '213333':
            encontreinsumo = 0
            for p in pedidos:
                letra = p.item_cont[0]
                if letra == 'A':

                    if p.item_cont == 'A 25' or p.item_cont == 'A 01':
                        encontreinsumo = 1

            if encontreinsumo == 0:
                nov = insumo + ', insumo sin actividad'
                crear_novedad(pedido, nov)

        if insumo == '211673':
            encontreinsumo = 0
            for p in pedidos:
                letra = p.item_cont[0]
                if letra == 'A':

                    if p.item_cont == 'A 23':
                        encontreinsumo = 1

            if encontreinsumo == 0:
                nov = insumo + ', insumo sin actividad'
                crear_novedad(pedido, nov)
                
        if insumo == '335931':
            encontreinsumo = 0
            for p in pedidos:
                letra = p.item_cont[0]
                if letra == 'A':

                    if p.item_cont == 'A 23':
                        encontreinsumo = 1

            if encontreinsumo == 0:
                nov = insumo + ', insumo sin actividad'
                crear_novedad(pedido, nov)

        if insumo == '200410' or insumo == '200411':
            encontreinsumo = 0
            for p in pedidos:
                letra = p.item_cont[0]
                if letra == 'A':

                    if p.item_cont == 'A 06' or p.item_cont == 'A 01' or p.item_cont == 'A 39' or p.item_cont == 'A 41' or p.item_cont == 'A 42':
                        encontreinsumo = 1

            if encontreinsumo == 0:
                nov = insumo + ', insumo sin actividad'
                crear_novedad(pedido, nov)

        if insumo == '211829':
            encontreinsumo = 0
            for p in pedidos:
                letra = p.item_cont[0]
                if letra == 'A':

                    if p.item_cont == 'A 02' or p.item_cont == 'A 27' or p.item_cont == 'A 01':
                        encontreinsumo = 1

            if encontreinsumo == 0:
                nov = insumo + ', insumo sin actividad'
                crear_novedad(pedido, nov)

        if insumo == '210948' or insumo == '210949':

            encontreinsumo = 0

            for p in pedidos:
                letra = p.item_cont[0]
                if letra == 'C':
                    encontreinsumo = 1
                if letra == 'A':

                    if p.item_cont == 'A 04' or p.item_cont == 'A 01' or p.actividad == 'ALEGA' or p.actividad == 'ALECA' or p.actividad == 'ALEGN' or p.actividad == 'ACAMN':
                        encontreinsumo = 1

            if encontreinsumo == 0:
                nov = insumo + ', insumo sin actividad'
                crear_novedad(pedido, nov)

    except:
        pass

def busqueda_item(pedido, item, item2, novedad):

    if item == 'A 03':
        try:
            busquedad_A03 = Acta.objects.filter(
                pedido=pedido.pedido).filter(item_cont=item).count()
            busquedad_A27 = Acta.objects.filter(
                pedido=pedido.pedido).filter(item_cont=item2).count()

            if busquedad_A03 <= 0:
                novedad = novedad+" "+str(item)+"=0."
                crear_novedad(pedido, novedad)
            if busquedad_A27 > 0:
                novedad = novedad+" incompatible con "+str(item2)+">0."
                crear_novedad(pedido, novedad)
                
        except:
            pass

    elif item == '210948':
        try:
            busquedad_210949 = Acta.objects.filter(
                pedido=pedido.pedido).filter(item_cont=item).count()
            busquedad_210948 = Acta.objects.filter(
                pedido=pedido.pedido).filter(item_cont='210949').count()

            if busquedad_210949 == 0 and busquedad_210948 == 0:
                novedad = novedad+" "+str(item)+"=0, 210949=0"
                crear_novedad(pedido, novedad)
        except:
            pass
    elif pedido.item_cont == 'A 42':
        try:
            busquedad_200410 = Acta.objects.filter(
                pedido=pedido.pedido).filter(item_cont=item).count()
            busquedad_200411 = Acta.objects.filter(
                pedido=pedido.pedido).filter(item_cont='200411').count()
            busquedad_200316 = Acta.objects.filter(
                pedido=pedido.pedido).filter(item_cont='200316').count()

            if busquedad_200410 == 0 and busquedad_200411 == 0 and busquedad_200316 == 0:
                novedad = novedad+" "+str(item)+"=0, 200411=0, 200316=0 "
                crear_novedad(pedido, novedad)
        except:
            pass

    elif (pedido.item_cont == 'A 10' or pedido.item_cont == 'A 11') and item == '200410':
        try:
            busquedad_200410 = Acta.objects.filter(
                pedido=pedido.pedido).filter(item_cont=item).count()
            busquedad_200411 = Acta.objects.filter(
                pedido=pedido.pedido).filter(item_cont='200411').count()

            busquedad_a03 = Acta.objects.filter(
                pedido=pedido.pedido).filter(item_cont='A 03').count()
            busquedad_a28 = Acta.objects.filter(
                pedido=pedido.pedido).filter(item_cont='A 28').count()
            busquedad_a29 = Acta.objects.filter(
                pedido=pedido.pedido).filter(item_cont='A 29').count()

            if busquedad_a03 != 0 and busquedad_a28 != 0 and busquedad_a29 != 0:

                if busquedad_200410 == 0 and busquedad_200411 == 0:
                    novedad = novedad + " 200410=0, 200411=0."
                    crear_novedad(pedido, novedad)
                    
            elif busquedad_200410<1 and busquedad_200411<1:
                novedad = novedad + " 200410=0, 200411=0."
                crear_novedad(pedido, novedad)
        except:
            pass

    elif item2 != 0:

        try:
            busquedad_1 = Acta.objects.filter(
                pedido=pedido.pedido).filter(item_cont=item).count()
            busquedad_2 = Acta.objects.filter(
                pedido=pedido.pedido).filter(item_cont=item2).count()

            if busquedad_1 == 0 and busquedad_2 == 0:
                novedad = novedad+" "+str(item)+"=0, " + str(item2)+"=0, "
                crear_novedad(pedido, novedad)
        except:
            pass
    else:
        try:
            busquedad = Acta.objects.filter(
                pedido=pedido.pedido).filter(item_cont=item).count()
            if busquedad == 0:
                novedad = novedad+" "+str(item)+"=0"
                crear_novedad(pedido, novedad)
        except:
            pass

def calculo_incompatible_A01(pedido, novedad):
    try:
        pedidos = Acta.objects.filter(pedido=pedido.pedido)
        cont = 1
        for p in pedidos:
            letra = p.item_cont[0]
            if letra == 'A':
                if p.item_cont != 'A 01' and p.item_cont != 'A 04' and p.item_cont != 'A 23' and p.item_cont != 'A 05':
                    nov = 'A 01=1, ' + p.item_cont + ">0. Incompatibles."
                    crear_novedad(p, nov)
    except:
        pass

def calculo_incompatible_A27(pedido, novedad):
    try:
        pedidos = Acta.objects.filter(pedido=pedido.pedido)
        cont = 1
        for p in pedidos:
            letra = p.item_cont[0]
            if letra == 'A':
                if p.item_cont == 'A 02' or p.item_cont == 'A 44' or p.item_cont == 'A 03':
                    nov = 'A 27=1, ' + p.item_cont + ">0. Incompatibles."
                    crear_novedad(p, nov)
    except:
        pass

def calculo_incompatible_A44(pedido, novedad):
    try:
        pedidos = Acta.objects.filter(pedido=pedido.pedido)
        cont = 1
        for p in pedidos:
            letra = p.item_cont[0]
            if letra == 'A':
                if p.item_cont == 'A 27' or p.item_cont == 'A 03':
                    nov = 'A 44=1, ' + p.item_cont + ">0. Incompatibles."
                    crear_novedad(p, nov)
    except:
        pass

def calculo_incompatible_A03(pedido, novedad):
    try:
        pedidos = Acta.objects.filter(pedido=pedido.pedido)
        cont = 1
        for p in pedidos:
            letra = p.item_cont[0]
            if letra == 'A':
                if p.item_cont == 'A 27' or p.item_cont == 'A 44':
                    nov = 'A 03=1, ' + p.item_cont + ">0. Incompatibles."
                    crear_novedad(p, nov)
    except:
        pass

def calculo_otros_incompatibles(pedido, item_comparar, comparado):
    try:
        pedidos = Acta.objects.filter(pedido=pedido.pedido)
        for p in pedidos:
            if p.item_cont == item_comparar:
                nov = str(comparado) + ' incompatible con ' + item_comparar
                crear_novedad(p, nov)
    except:
        pass

def crear_novedad(pedido, nov):
    # Validar si ya existe una novedad con ese pedido y novedad
    existe = Novedad_acta.objects.filter(pedido=pedido.pedido, novedad=nov).exists()
    if existe:
        return  # O puedes lanzar una excepción, según lo necesites

    # Si no existe, crear la novedad
    novedad = Novedad_acta()
    novedad.pedido = pedido.pedido
    novedad.actividad = pedido.actividad
    novedad.municipio = pedido.municipio
    novedad.pagina = pedido.pagina
    novedad.item = pedido.item_cont
    novedad.novedad = nov
    novedad.fecha = pedido.fecha_estado
    novedad.save()


def limpiar_novedades(request):
    CantidadItem.objects.all().delete()
    Novedad_acta.objects.all().delete()
    return redirect('novedades_acta')

def limpiar_acta(request):
    Acta.objects.all().delete()
    Novedad_acta.objects.all().delete()
    return redirect('home')

def calculo_enepre(pedido):
    if pedido.item_cont == 'A 03':
        try:
            novedad = "A 03=1, "
            busqueda_item(pedido, 'A 28', 0, novedad)
            busqueda_item(pedido, 'A 29', 0, novedad)
            busqueda_item(pedido, '219404', 0, novedad)
            busqueda_item(pedido, '325998',  0, novedad)
            calculo_incompatible_A03(pedido, novedad)
        except:
            pass

    if pedido.item_cont == 'A 27':
        try:
            novedad = "A 27=1, "
            busqueda_item(pedido, 'A 41', 0, novedad)
            busqueda_item(pedido, '200098', '200099', novedad)
            busqueda_item(pedido, '219404', 0, novedad)
            busqueda_item(pedido, 'A 10', 'A 11', novedad)
            calculo_incompatible_A27(pedido, novedad)
        except:
            pass

    if pedido.item_cont == 'A 44':
        try:
            novedad = "A 44=1, "
            busqueda_item(pedido, 'A 39', 0, novedad)
            busqueda_item(pedido, 'A 40', 'A 41', novedad)
            busqueda_item(pedido, '219404', 0, novedad)
            calculo_incompatible_A44(pedido, novedad)

            pedidos = Acta.objects.filter(pedido=pedido.pedido)

            encontre_A40 = 0
            encontre_A10 = 0
            encontre_A11 = 0
            encontre_riel = 0

            for p in pedidos:
                if p.item_cont == 'A 40':
                    encontre_A40 = 1
                if p.item_cont == 'A 10':
                    encontre_A10 = 1
                if p.item_cont == 'A 11':
                    encontre_A11 = 1
                if p.item_cont == '220683':
                    encontre_riel = 1

            if encontre_A40 == 0 and encontre_A10 == 0 and encontre_A11 == 0:
                novedad = novedad+"A 10=0, A 11=0"
                crear_novedad(pedido, novedad)
            if encontre_A40 == 1:
                if encontre_A10 == 1 or encontre_A11 == 1:
                    novedad = novedad+"A 40 incompatible con A 10 y A 11"
                    crear_novedad(pedido, novedad)
                if encontre_riel == 1:
                    nov = "A 40 excluye riel (220683)."
                    crear_novedad(pedido, nov)

        except:
            pass

def novedades_acta(request):
    novedades = {}
    try:
        novedades = Novedad_acta.objects.all()
    except:
        pass
    return render(request, "analisis.html", {'novedades': novedades})

def verificar_cable_acta(pedido, cable1, cable2, medidor):

    pedido1 = Acta.objects.filter(pedido=pedido).filter(item_cont=cable1)

    if len(pedido1) > 0:
        for p in pedido1:
            if float(p.cantidad) > 1:
                novedad = "Medidor: " + \
                    str(medidor)+" con cable " + \
                    str(p.suminis) + "= " + str(p.cantidad)

                pedido2 = Acta.objects.filter(
                    pedido=p).filter(item_cont=cable2)

                if len(pedido2) > 0:
                    for p in pedido2:
                        if float(p.cantidad) > 0:
                            novedad += " - " + \
                                str(p.suminis) + " : " + str(p.cantidad)

                crear_novedad(p, novedad)

def verificar_A03_MEDIDOR():
    # Obtén todos los registros donde item_cont es "A 03"
    actas_a03 = Acta.objects.filter(item_cont="A 03")

    for acta in actas_a03:
        # Verifica si hay otros registros con el mismo pedido y item_cont no igual a "A 03"
        otros_actas = Acta.objects.filter(
            Q(pedido=acta.pedido),
            ~Q(item_cont="A 03"),
            Q(item_cont="200092") | Q(item_cont="200093") | Q(item_cont="200098") | Q(item_cont="200099"),
            cantidad__gt=0
        )

        # Si no hay otros registros que cumplan las condiciones, llama a la función para crear una novedad
        if not otros_actas.exists():
            crear_novedad(acta, "A 03 sin medidor")
            
def filtrar_aejdo_sin_interna():
    # Filtrar pedidos con actividad AEJDO y item_cont que comience con B (incluir repetidos)
    pedidos_filtrados = Acta.objects.filter(
        actividad='AEJDO',
        item_cont__startswith='B'
    )   
    
    
    # Eliminar pedidos que tengan item_cont B 03
    pedidos_sin_b03 = pedidos_filtrados.exclude(item_cont='B 03')


    # Eliminar duplicados
    pedidos_finales = pedidos_sin_b03.values('pedido').distinct()
    


"""Nuevo analisis revision acta"""

from django.shortcuts import render
from django.contrib import messages
from analisis_acta.models import Acta, VariableAnalisis

def analisis_revision_acta(request):
    variable_region = VariableAnalisis.objects.first()  # Obtener la primera variable
    
    if variable_region:  # Si existe una variable, eliminar registros en Acta
        Acta.objects.exclude(subz=variable_region.region).delete()
    else:  
        messages.warning(request, "No existen variables de contrato registradas.")
        return redirect("index_admin")  # Ajusta el template
    