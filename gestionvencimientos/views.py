from asyncio.windows_events import NULL
from datetime import datetime, timedelta
from email.policy import HTTP
from django.http.response import HttpResponse, HttpResponseRedirect
import holidays_co
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from gestionvencimientos.models import Actividad, Actividad_epm, Ans, Encargado, Vencido
from material_oficiales.models import Despacho, Inicio, Liquidacion_acta_epm, Material_A_Buscar, Material_utilizado_perseo, Oficial, Reintegro, Stock
from medidores.models import NovedadMedidores, PedidoMedidores
from perseovsfenix.models import Guia, NumeroActa
from openpyxl import load_workbook
from django.urls import reverse
from django.http import JsonResponse
from django.shortcuts import render, redirect
from .models import Ans
from io import BytesIO

@login_required
def registrar_acta(request):
    if request.method == 'POST':
        numero = request.POST.get('numero')
        
        # Asegurarse de que siempre haya un solo registro
        acta_actual = NumeroActa.objects.first()
        
        if acta_actual:
            # Si ya existe un registro, actualízalo
            acta_actual.numero = numero
            acta_actual.save()
        else:
            # Si no existe, crea uno nuevo
            NumeroActa.objects.create(numero=numero)
        
        # Redirigir a la misma página o a otra página que se desee
        return redirect('home')  # Reemplaza con el nombre de la vista que corresponde
    
    return HttpResponse('Método no permitido', status=405)

@login_required
def index(request):
    return render(request,  "index.html")

def subir_acta_ans(request):
    try:
        if request.method == 'POST':
            # Verificar si el archivo fue subido correctamente
            if 'file' not in request.FILES:
                return JsonResponse({'status': 'error', 'message': 'No file provided'}, status=400)
             
            file = request.FILES['file']
            # Procesar el archivo en segundo plano (en este ejemplo, se hace síncronamente)
            process_excel(file)
            return HttpResponseRedirect(reverse('home'))  # Usamos HttpResponseRedirect en lugar de redirect
        
        # Si no es un POST, solo renderizamos la página
        return render(request, 'subir_extraccion_ans.html')
    
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


def process_excel(file):
    try:
        # Cargar el archivo en memoria usando BytesIO
        content = file.read()
        wb = load_workbook(filename=BytesIO(content))
        ws = wb.active  # Usamos la hoja activa por defecto
        row_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):  # Comenzamos desde la fila 2
            row_count += 1

            # Crear una nueva instancia del modelo Ans
            ans = Ans()

            # Asignar los valores de las filas del Excel a los campos del modelo Ans
            ans.Pedido = row[0]
            ans.Subped = row[1]
            ans.Soli = row[2]
            ans.Producto_id = row[3]
            ans.Tipo_Trabajo = row[4]
            ans.Tipo_Elemento_ID = row[5]
            ans.Fecha_Recibo = row[6]
            ans.Fecha_Ingreso_Sol = row[7]
            ans.Fecha_Concepto = row[8]
            ans.Fecha_Inicio_ANS = row[9]
            ans.Días_ANS = row[10] if row[10] is not None else 0.00
            ans.Estado = row[11]
            ans.Concepto = row[12]
            ans.Nombre_concepto = row[13]
            ans.ClienteID = row[14]
            ans.Nombre_Cliente = row[15]
            ans.Telefono = row[16]
            ans.Correo = row[17]
            ans.Direccion_Correspondencia = row[18]
            ans.Municipio_Correspondencia = row[19]
            ans.Telefono_Contacto = row[20]
            ans.Celular_Contacto = row[21]
            ans.Direccion = row[22]
            ans.Municipio = row[23]
            ans.Instalación = row[24]
            ans.Area_Operativa = row[25]
            ans.Subzona = row[26]
            ans.Area_Trabajo = row[27]
            ans.Ruta = row[28]
            ans.Coordenadax = row[29]
            ans.Coordenaday = row[30]
            ans.Actividad = row[31]
            ans.Equipo = row[32]
            ans.Nombre = row[33]
            ans.Fecha_Programación = row[34]
            ans.Num_Proyecto = row[35]
            ans.Tipo_Dirección = row[36]
            ans.Observación = row[37]
            ans.Observación_Solicitud = row[38]
            ans.Pedido_CRM = row[39]

            try:
                ans.save()
            except Exception as e:
                print(f"Error al guardar la fila {row_count}: {e}")

    except Exception as e:
        print(f"Error al procesar el archivo: {e}")

        
def calculo_dia_actutal():

    fecha_actual = datetime.now()
    dia = 0

    if fecha_actual.weekday() == 1:
        dia = 1

    if fecha_actual.weekday() == 2:
        dia = 2

    if fecha_actual.weekday() == 3:
        dia = 3

    if fecha_actual.weekday() == 4:
        dia = 4

    if fecha_actual.weekday() == 5:
        dia = 5

    if fecha_actual.weekday() == 6:
        dia = 6

    return dia


def calculo_dia_semana_2():

    fecha_actual = datetime.now()
    lunes = datetime.now()

    if fecha_actual.weekday() == 0:
        lunes = datetime.now()

    if fecha_actual.weekday() == 1:
        lunes = datetime.now()-timedelta(days=1)

    if fecha_actual.weekday() == 2:
        lunes = datetime.now()-timedelta(days=2)

    if fecha_actual.weekday() == 3:
        lunes = datetime.now()-timedelta(days=3)

    if fecha_actual.weekday() == 4:
        lunes = datetime.now()-timedelta(days=4)

    if fecha_actual.weekday() == 5:
        lunes = datetime.now()+timedelta(days=2)

    if fecha_actual.weekday() == 6:
        lunes = datetime.now()+timedelta(days=1)

    return lunes


def menu_pendientes(self):
    id_dia = calculo_dia_actutal()+1

    return redirect('pendientes', id_dia)


def limpiar_base(request):
    lista_ans = []
    aneses = Ans.objects.all().only('Subzona', 'Actividad')

    for ans in aneses:
        if ans.Subzona != "Uraba":
            ans.delete()

        elif ans.Actividad != "FSE" and ans.Actividad != "DSPRE" and ans.Actividad != "INFSM" and ans.Actividad != "ACREV" and ans.Actividad != "AEJDO" and ans.Actividad != "ARTER" and ans.Actividad != "DIPRE" and ans.Actividad != "INPRE" and ans.Actividad != "REEQU" and ans.Actividad != "APLIN" and ans.Actividad != "ALEGA" and ans.Actividad != "ALEGN" and ans.Actividad != "ALECA" and ans.Actividad != "ACAMN" and ans.Actividad != "AMRTR":
            ans.delete()          
            

    return redirect("gestionbd")


def calculo_pendientes(request, id_dia):

    lunes = calculo_dia_semana_2()
    encargados = Encargado.objects.all()

    if id_dia == 1:

        list_ans = busqueda_pendientes(lunes.strftime('%Y-%m-%d'))

        return render(request, "pendientes_lunes.html", {'id_dia': id_dia, 'encargados': encargados, 'aneses': list_ans, 'total': len(list_ans), 'fecha': lunes.strftime('%Y-%m-%d')})

    if id_dia == 2:

        list_ans = busqueda_pendientes(
            (lunes+timedelta(days=1)).strftime('%Y-%m-%d'))

        return render(request, "pendientes_lunes.html", {'id_dia': id_dia, 'encargados': encargados, 'aneses': list_ans, 'total': len(list_ans), 'fecha': (lunes+timedelta(days=1)).strftime('%Y-%m-%d')})

    if id_dia == 3:

        list_ans = busqueda_pendientes(
            (lunes+timedelta(days=2)).strftime('%Y-%m-%d'))

        return render(request, "pendientes_lunes.html", {'id_dia': id_dia, 'encargados': encargados, 'aneses': list_ans, 'total': len(list_ans), 'fecha': (lunes+timedelta(days=2)).strftime('%Y-%m-%d')})

    if id_dia == 4:

        list_ans = busqueda_pendientes(
            (lunes+timedelta(days=3)).strftime('%Y-%m-%d'))

        return render(request, "pendientes_lunes.html", {'id_dia': id_dia, 'encargados': encargados, 'aneses': list_ans, 'total': len(list_ans), 'fecha': (lunes+timedelta(days=3)).strftime('%Y-%m-%d')})

    if id_dia == 5:

        list_ans = busqueda_pendientes(
            (lunes+timedelta(days=4)).strftime('%Y-%m-%d'))

        return render(request, "pendientes_lunes.html", {'id_dia': id_dia, 'encargados': encargados, 'aneses': list_ans, 'total': len(list_ans), 'fecha': (lunes+timedelta(days=4)).strftime('%Y-%m-%d')})

    if id_dia == 6 or id_dia == 7:

        list_ans = busqueda_pendientes(
            (lunes).strftime('%Y-%m-%d'))

        return render(request, "pendientes_lunes.html", {'id_dia': id_dia, 'encargados': encargados, 'aneses': list_ans, 'total': len(list_ans), 'fecha': (lunes).strftime('%Y-%m-%d')})


def busqueda_vencidos(request):
    aneses = Ans.objects.filter(Estado="PENDI")
    list_ans = []
    for ans in aneses:
        if ans.Estado == "PENDI" and ans.Concepto == "PENDI":
            continue
        if ans.Estado == "PENDI" and ans.Concepto == "592":
            continue
        if ans.Tipo_Elemento_ID == "ENEGED":
            continue
        try:
            fecha_vence_ans = datetime.strptime(
                ans.fecha_vencimiento, "%Y-%m-%d %H:%M:%S")
            if fecha_vence_ans < datetime.today():
                if ans.estado_cierre == 0:
                    list_ans.append(ans)
        except Exception as e:
            print("An exception occurred in fecha vencimiento 2")
            print(repr(e))

    return render(request, "vencidos-todos.html", {"aneses": list_ans})


def busqueda_pendientes(fecha_vence_buscar):
    aneses = Ans.objects.filter(Estado="PENDI")
    list_ans = []
    for ans in aneses:
        if ans.Estado == "PENDI" and ans.Concepto == "PENDI":
            continue
        if ans.Estado == "PENDI" and ans.Concepto == "592":
            continue
        try:
            fecha_vence_ans = datetime.strptime(
                ans.fecha_vencimiento, "%Y-%m-%d %H:%M:%S")
            if fecha_vence_ans.strftime('%Y-%m-%d') == fecha_vence_buscar:
                if ans.estado_cierre == 0:
                    list_ans.append(ans)
        except Exception as e:
            print("An exception occurred in busqueda pendientes")
            print(repr(e))

    list_ans = cambiar_formato_fecha(list_ans)

    return list_ans


def cambiar_formato_fecha(fecha_a_cambiar):
    for l in fecha_a_cambiar:
        if l.fecha_vencimiento == None or l.fecha_vencimiento == '0':
            pass
        else:
            fecha = l.fecha_vencimiento.replace('-', '/')
            fecha_vencimiento = datetime.strptime(fecha, "%Y/%m/%d %H:%M:%S")
            l.fecha_vencimiento = fecha_vencimiento.strftime(
                "%d/%m/%Y %H:%M:%S")

    return fecha_a_cambiar


def cambiar_formato_fecha_epm(fecha_a_cambiar):
    for l in fecha_a_cambiar:
        if l.fecha_vencimiento == None or l.fecha_vencimiento == '0':
            pass
        else:
            fecha = l.fecha_vencimiento.replace('-', '/')
            fecha_vencimiento = datetime.strptime(fecha, "%Y/%m/%d %H:%M:%S")
            l.fecha_vencimiento = fecha_vencimiento.strftime(
                "%d/%m/%Y %H:%M:%S")

            fecha2 = l.fecha_vence_epm.replace('-', '/')
            fecha_vencimiento_epm = datetime.strptime(
                fecha2, "%Y/%m/%d %H:%M:%S")
            l.fecha_vence_epm = fecha_vencimiento_epm.strftime(
                "%d/%m/%Y %H:%M:%S")

            l.fecha_vence_sin_hora = fecha_vencimiento_epm.strftime("%d/%m/%Y")

    return fecha_a_cambiar


def eliminar_bd(request):
    Ans.objects.all().delete()

    return redirect('home')


def fechas(fecha_inic, dias):

    fecha_vencimiento = datetime. strptime(fecha_inic, '%Y-%m-%d %H:%M:%S')
    cont = 0

    while cont < dias:

        fecha_vencimiento = fecha_vencimiento+timedelta(days=1)

        if es_festivo_o_fin_de_semana(fecha_vencimiento):
            continue
        else:
            cont = cont+1

    return fecha_vencimiento


def es_festivo_o_fin_de_semana(fecha):
    if holidays_co.is_holiday_date(fecha):
        return True
    if fecha.weekday() == 5 or fecha.weekday() == 6:
        return True


def gestion_bd(request):

    lista_ans = []
    Ans.objects.filter(Estado="ANULA").delete()
    Ans.objects.filter(Estado="PENDI").filter(Concepto="PENDI").delete()

    anses = Ans.objects.all()

    for ans in anses:

        if ans.Actividad != "DSPRE" and ans.Actividad != "AEJDO" and ans.Actividad != "ACREV" and ans.Actividad != "ARTER" and ans.Actividad != "DIPRE" and ans.Actividad != "ACREV" and ans.Actividad != "INPRE" and ans.Actividad != "REEQU" and ans.Actividad != "APLIN" and ans.Actividad != "ALEGA" and ans.Actividad != "ALEGN" and ans.Actividad != "ALECA" and ans.Actividad != "ACAMN" and ans.Actividad != "AMRTR":
            ans.delete()

        if ans.Concepto != "405" and ans.Concepto != "INFSM" and ans.Concepto != "CEFSM" and ans.Concepto != "626" and ans.Concepto != "498" and ans.Concepto != "406" and ans.Concepto != "414" and ans.Concepto != "430" and ans.Concepto != "495" and ans.Concepto != "PENDI" and ans.Concepto != "FSE" and ans.Concepto != "PPRG" and ans.Concepto != "PROG":
            ans.delete()

    aneses = Ans.objects.all()

    for ans in aneses:

        try:
            if ans.Instalación[7:10] == "100":
                ans.Tipo_Dirección = "Urbano"
                ans.save()
            if ans.Instalación[7:10] == "200":
                ans.Tipo_Dirección = "Rural"
                ans.save()

            if ans.Fecha_Inicio_ANS == "":
                ans.Fecha_Inicio_ANS = ans.Fecha_Concepto
                ans.save()

            actividad = Actividad.objects.get(nombre=ans.Actividad)
            actividad_epm = Actividad_epm.objects.get(nombre=ans.Actividad)

            if ans.Tipo_Dirección == "Urbano":
                ans.dias_vencimiento = int(actividad.dias_urbano)
                ans.dias_vencimiento_epm = int(actividad_epm.dias_urbano)
            if ans.Tipo_Dirección == "Rural":
                ans.dias_vencimiento = int(actividad.dias_rural)
                ans.dias_vencimiento_epm = int(actividad_epm.dias_rural)

            fecha = fechas(ans.Fecha_Inicio_ANS, ans.dias_vencimiento)
            ans.fecha_vencimiento = fecha

            fecha_epm = fechas(ans.Fecha_Inicio_ANS, ans.dias_vencimiento_epm)
            ans.fecha_vence_epm = fecha_epm
            ans.fecha_vence_sin_hora = fecha.date().strftime("%d-%m-%Y")

            # inicio calculo vence epm

            # fin calculo vence epm

            ans.hora_vencimiento = fecha.time()

            ans.encargado = str(actividad.encargado)

            ans.save()

        except Exception as e:
            print("error gestion bd")
            print(e)

    return redirect('menu_pendientes')


def calculo_cable_medidores(pedido):
    try:
        mat200410 = Ans.objects.filter(item_cont='200410')
        mat200411 = Ans.objects.filter(item_cont='200411')
        if len(mat200410) > 0:
            print(mat200410)
        if len(mat200410) > 0:
            print(mat200410)
    except:
        pass


def cerrar_pedido(request, id_pedido):

    ans_cerrar = Ans.objects.get(id=id_pedido)

    ans_cerrar.estado_cierre = 1
    ans_cerrar.save()

    return redirect('menu_pendientes')


def calculo_next_week(request, id_dia):

    lunes = calculo_dia_semana_2()
    encargados = Encargado.objects.all()
    if id_dia == 10:

        list_ans = list_ans = busqueda_pendientes(
            (lunes+timedelta(days=7)).strftime('%Y-%m-%d'))

        return render(request, "pendientes_next_week.html", {'id_dia': id_dia, 'encargados': encargados, 'aneses': list_ans, 'total': len(list_ans), 'fecha': (lunes+timedelta(days=7)).strftime('%Y-%m-%d')})

    if id_dia == 20:

        list_ans = busqueda_pendientes(
            (lunes+timedelta(days=8)).strftime('%Y-%m-%d'))

        return render(request, "pendientes_next_week.html", {'id_dia': id_dia, 'encargados': encargados, 'aneses': list_ans, 'total': len(list_ans), 'fecha': (lunes+timedelta(days=8)).strftime('%Y-%m-%d')})

    if id_dia == 30:

        list_ans = busqueda_pendientes(
            (lunes+timedelta(days=9)).strftime('%Y-%m-%d'))

        return render(request, "pendientes_next_week.html", {'id_dia': id_dia, 'encargados': encargados, 'aneses': list_ans, 'total': len(list_ans), 'fecha': (lunes+timedelta(days=9)).strftime('%Y-%m-%d')})

    if id_dia == 40:

        list_ans = busqueda_pendientes(
            (lunes+timedelta(days=10)).strftime('%Y-%m-%d'))

        return render(request, "pendientes_next_week.html", {'id_dia': id_dia, 'encargados': encargados, 'aneses': list_ans, 'total': len(list_ans), 'fecha': (lunes+timedelta(days=10)).strftime('%Y-%m-%d')})

    if id_dia == 50:

        list_ans = busqueda_pendientes(
            (lunes+timedelta(days=11)).strftime('%Y-%m-%d'))

        return render(request, "pendientes_next_week.html", {'id_dia': id_dia, 'encargados': encargados, 'aneses': list_ans, 'total': len(list_ans), 'fecha': (lunes+timedelta(days=11)).strftime('%Y-%m-%d')})


def vencidos(request):
    aeneses = Vencido.objects.all()

    return render(request, "vencidos.html", {"aneses": aeneses})


def vencimientos_epm(request, inicio, final):
    fecha_inicio = inicio+" "+"00:00:00"
    fecha_final = final+" "+"23:59:59"
    aeneses = []
    ans = Ans.objects.all()
    for a in ans:
        if a.fecha_vence_epm > fecha_inicio and a.fecha_vence_epm < fecha_final:
            aeneses.append(a)
    aeneses = cambiar_formato_fecha_epm(aeneses)
    return render(request, "pendientes_epm.html", {"aneses": aeneses})

def vencimientos_contrato(request, inicio, final):
    fecha_inicio = inicio+" "+"00:00:00"
    fecha_final = final+" "+"23:59:59"
    aeneses = []
    ans = Ans.objects.all()
    for a in ans:
        if a.fecha_vencimiento > fecha_inicio and a.fecha_vencimiento < fecha_final:
            aeneses.append(a)
    aeneses = cambiar_formato_fecha_epm(aeneses)
    return render(request, "pendientes_epm.html", {"aneses": aeneses})



def pedidos_week(request, id_week):

    lunes = calculo_dia_semana_2()

    dia = 0

    if id_week == 2:
        dia = 7
        list_ans_lunes = busqueda_pendientes(
            (lunes+timedelta(days=dia)).strftime('%Y-%m-%d'))
    else:
        list_ans_lunes = busqueda_pendientes(lunes.strftime('%Y-%m-%d'))

    list_ans_martes = busqueda_pendientes(
        (lunes+timedelta(days=dia+1)).strftime('%Y-%m-%d'))
    list_ans_lunes.extend(list_ans_martes)

    list_ans_miercoles = busqueda_pendientes(
        (lunes+timedelta(days=dia+2)).strftime('%Y-%m-%d'))
    list_ans_miercoles.extend(list_ans_lunes)

    list_ans_jueves = busqueda_pendientes(
        (lunes+timedelta(days=dia+3)).strftime('%Y-%m-%d'))
    list_ans_miercoles.extend(list_ans_jueves)

    list_ans_viernes = busqueda_pendientes(
        (lunes+timedelta(days=dia+4)).strftime('%Y-%m-%d'))
    list_ans_miercoles.extend(list_ans_viernes)

    if id_week == 2:
        viernes = (lunes+timedelta(days=dia+4)).strftime('%d-%m-%Y')
        lunes = (lunes+timedelta(days=7)).strftime('%d-%m-%Y')
    else:
        viernes = (lunes+timedelta(days=dia+4)).strftime('%d-%m-%Y')
        lunes = lunes.strftime('%d-%m-%Y')

    lista_pedidos = list_ans_miercoles

    return render(request, "pedidos_week.html", {"aneses": lista_pedidos, "lunes": lunes, "viernes": viernes})


def otros_pedidos(request, cliente, apla, pendi):

    if cliente == 1 and apla == 0 and pendi == 0:
        vencidos = Ans.objects.filter(Estado="CLIEN")
        return render(request, "otros_pedidos.html", {"aneses": vencidos})

    if cliente == 0 and apla == 2 and pendi == 0:
        vencidos = Ans.objects.filter(Estado="APLAZ")
        return render(request, "otros_pedidos.html", {"aneses": vencidos})

    if cliente == 0 and apla == 0 and pendi == 3:
        vencidos = Ans.objects.filter(Estado="PENDI")
        return render(request, "otros_pedidos.html", {"aneses": vencidos})

    if cliente == 1 and apla == 2 and pendi == 0:
        vencidos = Ans.objects.filter(Q(Estado="CLIEN") | Q(Estado="APLAZ"))
        return render(request, "otros_pedidos.html", {"aneses": vencidos})

    if cliente == 0 and apla == 2 and pendi == 3:
        vencidos = Ans.objects.filter(Q(Estado="PENDI") | Q(Estado="APLAZ"))
        return render(request, "otros_pedidos.html", {"aneses": vencidos})

    if cliente == 1 and apla == 0 and pendi == 3:
        vencidos = Ans.objects.filter(Q(Estado="PENDI") | Q(Estado="CLIEN"))
        return render(request, "otros_pedidos.html", {"aneses": vencidos})

    if cliente == 1 and apla == 2 and pendi == 3:
        vencidos = Ans.objects.filter(
            Q(Estado="PENDI") | Q(Estado="CLIEN") | Q(Estado="APLAZ"))
        return render(request, "otros_pedidos.html", {"aneses": vencidos})

    return HttpResponse("Error en la consulta de otros pedidos")


def acrev(request):

    aeneses = Ans.objects.filter(Actividad="ACREV").filter(Q(Estado="PENDI") | Q(
        Concepto="406") | Q(Concepto="414") | Q(Concepto="495") | Q(Concepto="430"))
    aeneses = cambiar_formato_fecha(aeneses)
    return render(request, "acrev.html", {"aneses": aeneses})


def inconsistencias(request):

    aeneses1 = Ans.objects.filter(Concepto='FSE')
    aeneses2 = Ans.objects.filter(Concepto='INFSM')
    aenese = []
    print(len(aeneses2))
    for a in aeneses1:
        aenese.append(a)
    for ae in aeneses2:
        aenese.append(ae)

    aeneses = cambiar_formato_fecha(aenese)
    return render(request, "inconsistencias.html", {"aneses": aeneses})


def calculo_last_week(request, id_dia):

    lunes = calculo_dia_semana_2()
    encargados = Encargado.objects.all()
    if id_dia == 10:

        list_ans = list_ans = busqueda_pendientes(
            (lunes-timedelta(days=7)).strftime('%Y-%m-%d'))
        return render(request, "pendientes_last_week.html", {'id_dia': id_dia, 'encargados': encargados, 'aneses': list_ans, 'total': len(list_ans), 'fecha': (lunes+timedelta(days=7)).strftime('%Y-%m-%d')})

    if id_dia == 20:

        list_ans = busqueda_pendientes(
            (lunes-timedelta(days=6)).strftime('%Y-%m-%d'))
        return render(request, "pendientes_last_week.html", {'id_dia': id_dia, 'encargados': encargados, 'aneses': list_ans, 'total': len(list_ans), 'fecha': (lunes+timedelta(days=8)).strftime('%Y-%m-%d')})

    if id_dia == 30:

        list_ans = busqueda_pendientes(
            (lunes-timedelta(days=5)).strftime('%Y-%m-%d'))
        return render(request, "pendientes_last_week.html", {'id_dia': id_dia, 'encargados': encargados, 'aneses': list_ans, 'total': len(list_ans), 'fecha': (lunes+timedelta(days=9)).strftime('%Y-%m-%d')})

    if id_dia == 40:

        list_ans = busqueda_pendientes(
            (lunes-timedelta(days=4)).strftime('%Y-%m-%d'))
        return render(request, "pendientes_last_week.html", {'id_dia': id_dia, 'encargados': encargados, 'aneses': list_ans, 'total': len(list_ans), 'fecha': (lunes+timedelta(days=10)).strftime('%Y-%m-%d')})

    if id_dia == 50:

        list_ans = busqueda_pendientes(
            (lunes-timedelta(days=3)).strftime('%Y-%m-%d'))
        return render(request, "pendientes_last_week.html", {'id_dia': id_dia, 'encargados': encargados, 'aneses': list_ans, 'total': len(list_ans), 'fecha': (lunes+timedelta(days=11)).strftime('%Y-%m-%d')})
# aqui


# CODIGO INVENTARIO BODEGA

def gestionar_acta_perseo_inventario(request):

    pedidos_perseo = Material_utilizado_perseo.objects.all()
    pedidos_epm = Liquidacion_acta_epm.objects.all()

    for p in pedidos_perseo:
        try:

            codigo = p.codigo
            codigo_ultima_letra = codigo[-1]
            if codigo_ultima_letra == 'A' or codigo_ultima_letra == 'P':
                p.codigo = str(codigo[:-1])
                p.save()
        except:
            pass
        p.conc_pedido_codigo = str(p.pedido)+"-"+str(p.codigo)
        p.save()
    cont = 1
    for p in pedidos_epm:

        try:
            codigo = p.item_cont
            codigo_ultima_letra = codigo[-1]
            if codigo_ultima_letra == 'A' or codigo_ultima_letra == 'P':
                p.item_cont = str(codigo[:-1])
                p.save()
        except:
            pass

        try:
            nombre_cambio_codigo = Guia.objects.filter(
                nombre_fenix=p.item_cont)
            for n in nombre_cambio_codigo:
                p.item_cont = n.nombre_perseo
                p.save()

        except:
            pass

        p.conc_pedido_codigo = str(p.pedido)+"-"+str(p.item_cont)
        p.save()

        pedido_a_modificar = Material_utilizado_perseo.objects.filter(pedido=p)

        for ped in pedido_a_modificar:
            p.encargado = ped.instalador
            p.save()
            break

    return render(request,  "index.html")


def calculo_inventario_por_oficial(request):
    cont = 0
    for oficial in Oficial.objects.all():
        for material_a_buscar in Material_A_Buscar.objects.all():

            inicio = 0
            despachado = 0
            epm = 0
            diferencia = 0
            reintegro = 0

            cantidad_inicial_inicio = Inicio.objects.filter(
                encargado=oficial).filter(codigo=material_a_buscar)

            for cant_inicio in cantidad_inicial_inicio:
                inicio += float(cant_inicio.cantidad)

            cantidad_despacho = Despacho.objects.filter(
                encargado=oficial).filter(codigo=material_a_buscar)
            for cant_des in cantidad_despacho:
                despachado += float(cant_des.cantidad)

            cantidad_reintegro = Reintegro.objects.filter(
                encargado=oficial).filter(codigo=material_a_buscar)
            for cant_rei in cantidad_reintegro:
                cont += 1
                reintegro += float(cant_rei.cantidad)

            cantidad_usado_en_campo = Liquidacion_acta_epm.objects.filter(
                encargado=oficial).filter(item_cont=material_a_buscar)
            for cant_epm in cantidad_usado_en_campo:
                epm += float(cant_epm.cantidad)

            diferencia = int(inicio)-int(reintegro)+int(despachado)-int(epm)

            stock = Stock()
            stock.encargado = oficial.nombre
            stock.codigo = material_a_buscar.nombre
            stock.inicio = inicio
            stock.despachado = despachado
            stock.reintegrado = reintegro
            stock.epm = epm
            stock.diferencia = diferencia
            stock.save()

    return render(request,  "index.html")


def reiniciar_bd_oficiales(request):
    Liquidacion_acta_epm.objects.all().delete()
    Inicio.objects.all().delete()
    Despacho.objects.all().delete()
    Stock.objects.all().delete()
    Reintegro.objects.all().delete()
    Material_utilizado_perseo.objects.all().delete()
    Material_A_Buscar.objects.all().delete()
    return render(request,  "index.html")

# CODIGO MEDIDORES VS CABLEADO USADO



def gestion_medidores():
    pedidos = PedidoMedidores.objects.filter(suminis='200092') | PedidoMedidores.objects.filter(
        suminis='200093') | PedidoMedidores.objects.filter(suminis='200098')

    for ped in pedidos:
        medidor = ped.suminis
        if ped.suminis == '200092' or ped.suminis == '200098':
            verificar_cable(ped.pedido, '200411', '200410', medidor)

        if ped.suminis == '200093':
            verificar_cable(ped, '200410', '200411', medidor)


def verificar_cable(pedido, cable1, cable2, medidor):

    pedido1 = PedidoMedidores.objects.filter(
        pedido=pedido).filter(item_cont=cable1)

    if len(pedido1) > 0:
        for p in pedido1:
            if float(p.cantidad) > 1:
                novedad = "Medidor: " + \
                    str(medidor)+" con cable " + \
                    str(p.suminis) + " : " + str(p.cantidad)

                pedido2 = PedidoMedidores.objects.filter(
                    pedido=p).filter(item_cont=cable2)

                if len(pedido2) > 0:
                    for p in pedido2:
                        if float(p.cantidad) > 0:
                            novedad += " - " + \
                                str(p.suminis) + " : " + str(p.cantidad)

                crearNovedad(p, novedad)


def crearNovedad(pedido, novedad):
    nov = NovedadMedidores()
    nov.pedido = pedido
    nov.novedad = novedad
    nov.save()


def reiniciar_medidores(request):
    PedidoMedidores.objects.all().delete()
    NovedadMedidores.objects.all().delete()
    return render(request,  "index.html")


