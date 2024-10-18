from django.shortcuts import render, redirect
from django.contrib import messages
import openpyxl
from django.db.models import Sum
from .models import Pedido, ActaB, ComparacionPedido, MaterialSeleccionado

def formulario_subir_pedido(request):
    return render(request, 'formulario_subir_perseo.html')  # Asegúrate de usar el nombre correcto del template

def formulario_subir_acta(request):
    return render(request, 'formulario_subir_acta.html')  # Asegúrate de usar el nombre correcto del template

def subir_material_mejia(request):
    if request.method == 'POST':
        file = request.FILES.get('file')
        
        if not file:
            messages.error(request, 'Por favor selecciona un archivo.')
            return redirect('ver_material_mejia')

        # Verificar que el archivo sea de tipo Excel
        if not file.name.endswith('.xlsx'):
            messages.error(request, 'El archivo debe ser un archivo Excel (.xlsx).')
            return redirect('ver_material_mejia')

        try:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active  # Seleccionar la hoja activa del archivo Excel

            # Obtener todos los códigos y guías desde MaterialSeleccionado
            seleccionados = MaterialSeleccionado.objects.values('codigo', 'guia')

            # Crear un diccionario para mapear codigo a guia
            codigo_to_guia = {s['codigo']: s['guia'] for s in seleccionados}

            # Empezar desde la segunda fila (ignorar el encabezado)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                codigo = row[18]  # Columna S (índice 18)
                
                # Convertir el código a string en caso de que sea numérico
                codigo = str(codigo)
                
                # Verificar si el código está en MaterialSeleccionado
                if codigo in codigo_to_guia:
                    pedido = row[3]   # Columna D (índice 3)
                    actividad = row[11]  # Columna L (índice 11)
                    instalador = row[12]  # Columna M (índice 12)
                    cantidad = row[20]  # Columna U (índice 20)
                    fecha = row[25]  # Columna Z (índice 25)
                    acta = row[26]  # Columna AA (índice 26)

                    # Obtener la guía correspondiente del diccionario
                    guia = codigo_to_guia[codigo]

                    # Crear la instancia del modelo Pedido y guardarla con la guía
                    Pedido.objects.create(
                        pedido=pedido,
                        actividad=actividad,
                        instalador=instalador,
                        codigo=codigo,
                        guia=guia,  # Asignar la guía correspondiente
                        cantidad=cantidad,
                        fecha=fecha,
                        acta=acta
                    )
            
            messages.success(request, 'Archivo subido y procesado correctamente.')
            return redirect('ver_material_mejia')

        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
            print(e)
            return redirect('ver_material_mejia')

    return redirect('ver_material_mejia')

def ver_material_mejia(request):
    pedidos = Pedido.objects.all()  # Obtén todos los pedidos desde la base de datos
    return render(request, 'material_mejia.html', {'pedidos': pedidos}) 

def ver_material_acta(request):
    pedidos = ActaB.objects.all()  # Obtén todos los pedidos desde la base de datos
    return render(request, 'material_acta.html', {'actas': pedidos}) 

def subir_material_acta(request):
    if request.method == 'POST':
        file = request.FILES.get('file')

        if not file:
            messages.error(request, 'Por favor selecciona un archivo.')
            return redirect('ver_material_acta')

        # Verificar que el archivo sea de tipo Excel
        if not file.name.endswith('.xlsx'):
            messages.error(request, 'El archivo debe ser un archivo Excel (.xlsx).')
            return redirect('ver_material_acta')
        
        try:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active  # Seleccionar la hoja activa del archivo Excel
            
            # Obtener todos los valores del campo 'guia' del modelo MaterialSeleccionado
            guias_seleccionadas = MaterialSeleccionado.objects.values_list('guia', flat=True)

            # Empezar desde la segunda fila (ignorar el encabezado)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                
                
                pedido = row[0]  # Columna A (índice 0)

                actividad = row[7]  # Columna H (índice 7)
                
                # Columna S (índice 18) o R (índice 17) si S está vacía
                codigo = row[18] if row[18] else row[17]

                # Convertir el código a string en caso de que sea numérico
                codigo = str(codigo)
                
                # Si el código termina en 'A' o 'P', eliminar el último carácter
                if codigo and codigo[-1] in ['A', 'P']:
                    codigo = codigo[:-1]

                cantidad = row[20]  # Columna U (índice 20)

                # Verificar si el código está en el campo 'guia' del modelo MaterialSeleccionado                
                if codigo in guias_seleccionadas:
                    # Crear la instancia del modelo Acta y guardarla
                    ActaB.objects.create(
                        pedido=pedido,
                        actividad=actividad,
                        codigo=codigo,
                        cantidad=float(cantidad)
                    )

            messages.success(request, 'Archivo subido y procesado correctamente.')
            return redirect('ver_material_acta')

        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
            print(str(e))
            return redirect('ver_material_acta')

    return redirect('ver_material_acta')

def reiniciar_actas(request):
    Pedido.objects.all().delete()
    ActaB.objects.all().delete()
    ComparacionPedido.objects.all().delete()
    return redirect('lista_comparaciones')

def reiniciar_novedades(request):
    ComparacionPedido.objects.all().delete()
    return redirect('lista_comparaciones') 

def comparar_pedidos(request):
    
    # Verificar si hay códigos en ActaB que no están en Pedido
    verificar_codigos_actab_sin_pedido(request)
    
    # Obtener las sumas de materiales del modelo Pedido, incluyendo todos los campos necesarios
    pedidos = Pedido.objects.values(
        'pedido', 'codigo', 'guia', 'fecha', 'actividad', 'instalador'
    ).annotate(suma_pedido=Sum('cantidad'))

    # Obtener las sumas de materiales del modelo ActaB
    actas = ActaB.objects.values('pedido', 'codigo').annotate(suma_acta=Sum('cantidad'))

    # Lista para controlar los pedidos y guías ya procesados
    pedidos_procesados = set()

    # Iterar sobre los pedidos y buscar la guía correspondiente al código en ActaB
    for pedido in pedidos:
        # Crear un identificador único para cada combinación de pedido y guía (no el código)
        pedido_guia = (pedido['pedido'], pedido['guia'])

        # Si ya hemos procesado este pedido y guía, saltamos esta iteración
        if pedido_guia in pedidos_procesados:
            continue

        # Obtener la suma de todas las cantidades para el mismo pedido y código
        suma_material_pedido_total = Pedido.objects.filter(pedido=pedido['pedido'], guia=pedido['guia']).aggregate(total_pedido=Sum('cantidad'))['total_pedido']

        # Usamos la guía del pedido para encontrar el registro correspondiente en ActaB
        acta = next((a for a in actas if a['pedido'] == pedido['pedido'] and a['codigo'] == pedido['guia']), None)

        suma_material_acta = acta['suma_acta'] if acta else 0
        suma_material_pedido = suma_material_pedido_total if suma_material_pedido_total is not None else 0

        # Solo guardamos los registros donde las sumas no coinciden
        if suma_material_acta != suma_material_pedido:
            # Asegúrate de que 'fecha' y demás campos existen en el diccionario 'pedido'
            fecha = pedido.get('fecha', None)
            actividad = pedido.get('actividad', '')
            instalador = pedido.get('instalador', '')

            # Establecer la observación si las cantidades no coinciden
            observacion = 'Cantidad no coincide' if suma_material_acta != suma_material_pedido else ''

            # Calcular la diferencia
            diferencia = suma_material_acta - suma_material_pedido

            # Crear una instancia del modelo ComparacionPedido
            ComparacionPedido.objects.create(
                pedido=pedido['pedido'],
                codigo=pedido['codigo'],
                suma_material_pedido=suma_material_pedido,
                suma_material_acta=suma_material_acta,
                diferencia=diferencia,
                actividad=actividad,
                fecha=fecha,
                instalador=instalador,
                observacion=observacion
            )

            # Añadir este pedido y guía a la lista de procesados
            pedidos_procesados.add(pedido_guia)

    return redirect('lista_comparaciones')

def verificar_codigos_actab_sin_pedido(request):
    # Obtener todos los códigos y pedidos del modelo Pedido, incluyendo la actividad e instalador
    pedidos = Pedido.objects.values('pedido', 'codigo', 'actividad', 'instalador', 'fecha')

    # Convertir los pedidos y códigos en un conjunto para facilitar la búsqueda
    pedidos_codigos = set((p['pedido'], p['codigo']) for p in pedidos)

    # Obtener las guías del modelo MaterialSeleccionado que vinculan los códigos de Pedido y ActaB
    material_seleccionado = MaterialSeleccionado.objects.values('codigo', 'guia')

    # Crear un diccionario de 'guia' a 'codigo' para asociar código en Pedido con la guía en ActaB
    guia_to_codigo = {ms['guia']: ms['codigo'] for ms in material_seleccionado}

    # Obtener todos los códigos y pedidos del modelo ActaB
    actas = ActaB.objects.values('pedido', 'codigo', 'cantidad', 'actividad')

    # Iterar sobre los códigos de ActaB
    for acta in actas:
        
        # Verificar si el código en ActaB (que corresponde a la 'guía') no está asociado a un código en Pedido
        codigo_pedido = guia_to_codigo.get(acta['codigo'], None)

        # Si no hay un código en Pedido asociado a la guía (código en ActaB), buscamos actividad e instalador
        if not codigo_pedido or (acta['pedido'], codigo_pedido) not in pedidos_codigos:
            print(acta)
            # Si el código no está en Pedido, creamos un registro en ComparacionPedido
            ComparacionPedido.objects.create(
                pedido=acta['pedido'],
                codigo=acta['codigo'],  # Mostramos la guía (código en ActaB)
                suma_material_pedido=0,  # No hay datos en Pedido, así que la suma es 0
                suma_material_acta=acta['cantidad'],  # Usar el valor de ActaB
                diferencia=acta['cantidad'],  # Diferencia será igual a la suma de ActaB
                actividad=acta['actividad'],  # Actividad del pedido
                fecha=None,  # Fecha del pedido
                instalador='No disponible',  # Instalador del pedido
                observacion=f'Código {acta["codigo"]} no aparece en Perseo'
            )

    messages.success(request, 'Verificación completada. Registros creados para códigos que no aparecen en Perseo.')
    return redirect('lista_comparaciones')

def lista_comparaciones(request):
    # Obtener todas las comparaciones guardadas en la base de datos
    comparaciones = ComparacionPedido.objects.all()

    # Crear una lista para almacenar los datos con la 'guía' de MaterialSeleccionado
    comparaciones_con_guia = []

    # Obtener la relación 'codigo' y 'guia' de MaterialSeleccionado
    seleccionados = MaterialSeleccionado.objects.values('codigo', 'guia')

    # Crear un diccionario de 'codigo' a 'guia'
    codigo_to_guia = {s['codigo']: s['guia'] for s in seleccionados}

    # Iterar sobre las comparaciones y buscar la 'guia' correspondiente al 'codigo'
    for comparacion in comparaciones:
        guia = codigo_to_guia.get(comparacion.codigo, comparacion.codigo)  # Si no hay guía, usar el código como fallback
        comparaciones_con_guia.append({
            'pedido': comparacion.pedido,
            'actividad': comparacion.actividad,
            'codigo': guia,  # Mostrar la guía en lugar del código
            'suma_material_pedido': comparacion.suma_material_pedido,
            'suma_material_acta': comparacion.suma_material_acta,
            'diferencia': comparacion.diferencia,
            'instalador': comparacion.instalador,
            'observacion': comparacion.observacion,
            'fecha': comparacion.fecha,
        })

    return render(request, 'lista_comparaciones.html', {'comparaciones': comparaciones_con_guia})

