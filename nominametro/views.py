import pandas as pd
import io
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from openpyxl import load_workbook
from django.db.models import Sum
from nominametro.models import Mejia, Novedad_nomina, plantilla, prenomina, Concepto

from django.http import JsonResponse
from openpyxl import load_workbook

def subirnominametro(request):
    try:
        if request.method == 'POST':
            file = request.FILES['file']
            request.session['progress'] = 0  # Inicializa progreso
            process_excel(file, request)
            return JsonResponse({"message": "Archivo subido exitosamente"}, status=200)
        return render(request, 'subirarchivo.html')    
    except Exception as e:
        print(e)
        return JsonResponse({"error": str(e)}, status=500)

def process_excel(file, request):
    try:
        wb = load_workbook(file)
        ws = wb[wb.sheetnames[0]]
        total_rows = sum(1 for _ in ws.iter_rows()) - 1  # Ignorar encabezado
        processed_rows = 1
        
        for row in ws.iter_rows(min_row=2):
            nomina_empleado = prenomina()
            nomina_empleado.centro_de_costos = row[0].value
            nomina_empleado.nombre_del_centro_de_costos = row[1].value
            nomina_empleado.empleado = row[2].value
            nomina_empleado.nombre_del_empleado = row[3].value
            nomina_empleado.turno = row[4].value
            nomina_empleado.descripción_del_turno = row[5].value
            nomina_empleado.concepto = row[6].value
            
            try:
                nomina_empleado.conversor = buscar_conversor(row[6].value)
                tipo_ingreso = Concepto.objects.filter(concepto=row[6].value).first()
                nomina_empleado.tipo = tipo_ingreso.tipo
                nomina_empleado.factor = tipo_ingreso.factor
            except Exception as e:
                novedad = Novedad_nomina(empleado=nomina_empleado.empleado, novedad=str(e))
                novedad.save()

            nomina_empleado.nombre_del_concepto = row[7].value
            nomina_empleado.vinculación = row[8].value
            nomina_empleado.préstamo = row[9].value
            nomina_empleado.salario_básico_hora = row[10].value
            nomina_empleado.tiempo = row[11].value
            nomina_empleado.valor = row[12].value
            nomina_empleado.save()
            
            # Actualizar progreso
            processed_rows += 1
            request.session['progress'] = str(processed_rows) +" filas de " + str(total_rows)
            request.session.modified = True
            request.session.save()  # 🔥 Esto fuerza a guardar la sesión    
                    
    except Exception as e:
        print(e)

def get_progress(request):
    progress = request.session.get('progress', 0)
    return JsonResponse({'progress': progress})

def buscar_conversor(concepto):
    conversor = Concepto.objects.filter(concepto=concepto).first()
    conversor = conversor.conversor
    return conversor


def definir_fechas(request):
    return render(request, 'fecha_inicial_final.html')


def gestionar_prenomina(request):

    try:

        if request.method == 'POST':
            fecha_inicial = request.POST['fecha_inicial']
            fecha_final = request.POST['fecha_final']

            empleados = prenomina.objects.values_list('empleado').distinct()
            nit= Mejia.objects.all().first()
            

            for e in empleados.order_by('empleado'):

                empleado = prenomina.objects.filter(empleado=e[0]).first()

                emplea = e[0]

                nomina_empleado = plantilla()
                nomina_empleado.nit=str(nit)
                nomina_empleado.cedula = empleado.empleado
                nomina_empleado.nombre = empleado.nombre_del_empleado
                nomina_empleado.apellido = empleado.nombre_del_empleado
                nomina_empleado.periodo_fecha_inicial = fecha_inicial
                nomina_empleado.periodo_fecha_final = fecha_final
                nomina_empleado.valor_hora_ordin = empleado.salario_básico_hora

                horas_ordinarias = calculo_horas(emplea, 100)
                nomina_empleado.horas_ordinarias = horas_ordinarias

                on_0_35 = calculo_horas(emplea, 200)
                nomina_empleado.on_0_35 = on_0_35

                ed_1_25 = calculo_horas(emplea, 300)
                nomina_empleado.ed_1_25 = ed_1_25

                en_1_75 = calculo_horas(emplea, 400)
                nomina_empleado.en_1_75 = en_1_75

                fd_0_75 = calculo_horas(emplea, 500)
                nomina_empleado.fd_0_75 = fd_0_75

                fn_1_1 = calculo_horas(emplea, 600)
                nomina_empleado.fn_1_1 = fn_1_1

                efd_2 = calculo_horas(emplea, 700)
                nomina_empleado.efd_2 = efd_2

                efn_2_5 = calculo_horas(emplea, 800)
                nomina_empleado.efn_2_5 = efn_2_5

                d_o_f_d_1_75 = calculo_horas(emplea, 900)
                nomina_empleado.d_o_f_d_1_75 = d_o_f_d_1_75

                d_o_f_n_2_1 = calculo_horas(emplea, 1000)
                nomina_empleado.d_o_f_n_2_1 = d_o_f_n_2_1

                ausencias_remuneradas_hora = calculo_horas(emplea, 1100)
                nomina_empleado.ausencias_remuneradas_hora = ausencias_remuneradas_hora

                ausencias_no_remuneradas_hora = calculo_horas(emplea, 1200)
                nomina_empleado.ausencias_no_remuneradas_hora = ausencias_no_remuneradas_hora

                incapacidad_por_enfermedad_general_horas = calculo_horas(
                    emplea, 1300)
                nomina_empleado.incapacidad_por_enfermedad_general_horas = incapacidad_por_enfermedad_general_horas

                vr_auxilio_transporte_o_auxilio_de_conectividad = calculo_valor(
                    emplea, 1400)
                nomina_empleado.vr_auxilio_transporte_o_auxilio_de_conectividad = vr_auxilio_transporte_o_auxilio_de_conectividad

                otros_ingresos_no_prestacionales = calculo_valor(emplea, 1500)
                nomina_empleado.otros_ingresos_no_prestacionales = otros_ingresos_no_prestacionales

                otros_ingresos_prestacionales = calculo_valor(emplea, 1900)
                nomina_empleado.otros_ingresos_prestacionales = otros_ingresos_prestacionales

                nomina_empleado.total_devengado = calculo_devengado(empleado)

                nomina_empleado.deducción_retención_en_la_fuente = calculo_valor(
                    empleado, 1600)

                nomina_empleado.otras_deducciones = calculo_valor(emplea, 1700)

                nomina_empleado.deducciones_sgss = calculo_valor(emplea, 1800)

                nomina_empleado.incapacidad_por_accidente_laboral = calculo_horas(emplea, 2000)

                nomina_empleado.valor_pago_prestaciones = calculo_valor(emplea, 2100)

                nomina_empleado.neto_a_pagar = calculo_neto_a_pagar(emplea)

                nomina_empleado.definir_cargo()
                nomina_empleado.definir_salario()

                nomina_empleado.save()

            calculo_nombre_apellido()

    except Exception as e:
        print(e)

    return redirect('informe')


def reiniciar_prenomina(request):
    prenomina.objects.all().delete()
    plantilla.objects.all().delete()
    Novedad_nomina.objects.all().delete()
    return redirect('informe')


def calculo_horas(empleado, conversor):

    horas = prenomina.objects.filter(empleado=empleado).filter(
        conversor=conversor).aggregate(suma=Sum('tiempo'))

    if horas['suma'] is None:
        horas = 0
    else:
        horas = horas['suma']

    return horas


def calculo_valor(empleado, conversor):

    valor = prenomina.objects.filter(empleado=empleado).filter(
        conversor=conversor).aggregate(suma=Sum('valor'))

    if valor['suma'] is None:
        valor = 0
    else:
        valor = valor['suma']

    return valor


def calculo_devengado(empleado):
    valor = prenomina.objects.filter(empleado=empleado)\
        .filter(tipo='devengado').aggregate(devengado=Sum('valor'))
    devengado = valor['devengado']

    return devengado


def calculo_neto_a_pagar(empleado):
    valor = prenomina.objects.filter(empleado=empleado)\
        .filter(tipo='devengado').aggregate(devengado=Sum('valor'))

    deduccion = prenomina.objects.filter(empleado=empleado)\
                                 .filter(tipo='deduccion').aggregate(deduccion=Sum('valor'))

    if deduccion['deduccion'] is None:
        deduccion['deduccion'] = 0
    if valor['devengado'] is None:
        valor['devengado'] = 0

    neto_a_pagar = float(valor['devengado'])-float(deduccion['deduccion'])

    return neto_a_pagar


def informe(request):
    informe = plantilla.objects.all()
    return render(request, 'informe.html', {'informe': informe})


def calculo_nombre_apellido():

    informe = plantilla.objects.all()

    for i in informe:
        numero_palabras = len(i.nombre.split())

        if numero_palabras == 2:
            cont = 1
            for inf in i.nombre.split():
                if cont == 1:
                    i.apellido = inf
                else:
                    i.nombre = inf
                cont += 1

        if numero_palabras == 3:
            cont = 1
            apellido = ""
            for inf in i.nombre.split():
                if cont == 1:
                    apellido = inf
                if cont == 2:
                    apellido = apellido + " " + str(inf)
                    i.apellido = apellido
                if cont == 3:
                    i.nombre = inf
                cont += 1

        if numero_palabras == 4:
            cont = 1
            apellido = ""
            nombre = ""
            for inf in i.nombre.split():
                if cont == 1:
                    apellido = inf
                if cont == 2:
                    apellido = apellido + " " + str(inf)
                    i.apellido = apellido
                if cont == 3:
                    nombre = inf
                if cont == 4:
                    i.nombre = nombre + " " + str(inf)

                cont += 1

        if numero_palabras > 4:
            cont = 1
            apellido = ""
            nombre = ""
            for inf in i.nombre.split():
                if cont <= (numero_palabras-2) and cont < 3:
                    apellido += inf + " "
                if cont >= 3:
                    nombre += inf + " "
                cont += 1
            i.nombre = nombre
            i.apellido = apellido

        i.save()


def export_excel(request):
    registros = plantilla.objects.values_list()

    for r in registros:
        print(r)

    df = pd.DataFrame(registros, columns=['id','Nit Empresa Contratista o Subcontratita ','Cedula','Nombre','Apellido','Cargo','Salario Mensual Basico / Honorario mensual','Valor/hora ordinaria','Periodo Fecha Inicial (dd/mm/yyyy)','Periodo Fecha Final  (dd/mm/yyyy)','Horas Ordinarias (1,00)','Horas recargo nocturno  (0.35)','Horas extra diurna (1.25)','Hora extra nocturna (1.75)','Horas ordinarias festivas diurnas (0.75)','Horas recargo ordinaria festiva nocturna (1.1)','Horas ordinaria festiva nocturna (2.1)','Horas extras festivas diurnas  (2,0)','Horas extras festivas nocturas (2.5)','Horas domingo o festivo  diurno (1.75)','Horas ausencias remuneradas','Horas ausencias No remuneradas','Horas incapacidad por accidente laboral ','Horas incapacidad por enfermedad general','Valor auxilio Transporte o Auxilio de Conectividad','Valor otros Ingresos prestacionales','Valor otros Ingresos No prestacionales','Valor pago prestaciones (prima, cesantias, Int.cesantias, vacaciones)','Valor total Devengado','Valor deducción Retención en la Fuente','Valor otras Deducciones','Valor Deducciones  salud y pensión','VALOR NETO A PAGAR','OBSERVACIONES'])

    # Eliminar la columna "ID"
    df = df.drop('id', axis=1)

    # Crear objeto ExcelWriter
    writer = pd.ExcelWriter('registros.xlsx', engine='xlsxwriter')

    # Escribir DataFrame en archivo de Excel
    df.to_excel(writer, index=False, sheet_name='Registros')

    # Obtener objeto workbook y worksheet
    workbook = writer.book
    worksheet = writer.sheets['Registros']

    # Crear objeto header_format con estilo de encabezado
    header_format = workbook.add_format({
        'bold': True,
        'text_wrap': True,
        'valign': 'top',
        'fg_color': '#7AD400',
        'align': 'justify',
        'align': 'center',
        'valign': 'vcenter',
        'font_name': 'Trebuchet MS',
        'font_size': 9,
        'right': 1,
        'right_color':'white'
    })

    format = workbook.add_format({
        'num_format': '#,##0.00',
        'font_name': 'Trebuchet MS',
        'font_size': 8,
        'border':1,
        'border_color': '#00B050'
    })

    format1 = workbook.add_format({
        'font_name': 'Trebuchet MS',
        'font_size': 8,
         'border':1,
        'border_color': '#00B050'
    })

    format_fecha = workbook.add_format({
    'num_format': 'dd/mm/yyyy',  # Formato correcto para fechas
    'font_name': 'Trebuchet MS',
    'font_size': 8,
    'border': 1,
    'border_color': '#00B050',
    'align': 'center'
})

    format_horas = workbook.add_format({
        'font_name': 'Trebuchet MS',
        'font_size': 8,
        'num_format': '#,##0.00',
        'border':1,
        'border_color': '#00B050',
        'align': 'center'
    })

    format_firma = workbook.add_format({
        'font_name': 'Trebuchet MS',
        'num_format': '#.##0,00_-;#.##0,00_-;"-"??_-;_-@_-',
        'font_size': 8,
        'border':1,
        'border_color': '#00B050',
        'align': 'left'
    })

    worksheet.set_column(0, 0, 12, format1)
    worksheet.set_column(1, 1, 16, format1)
    worksheet.set_column(2, 2, 18, format1)
    worksheet.set_column(3, 3, 12, format1)
    worksheet.set_column(4, 5, 26, format)
    worksheet.set_column(6, 6, 18, format)
    worksheet.set_column(7, 8, 22, format_fecha)
    worksheet.set_column(9, 9, 21, format_horas)
    worksheet.set_column(10, 16, 12, format_horas)
    worksheet.set_column(17, 17, 16, format_horas)
    worksheet.set_column(18, 18, 15, format_horas)
    worksheet.set_column(19, 19, 23, format_horas)
    worksheet.set_column(20, 20, 27, format_horas)
    worksheet.set_column(21, 21, 28, format_horas)
    worksheet.set_column(22, 22, 25, format)
    worksheet.set_column(23, 23, 21, format)
    worksheet.set_column(24, 25, 18, format)
    worksheet.set_column(26, 26, 27, format)
    worksheet.set_column(27, 28, 22, format)
    worksheet.set_column(29, 29, 15, format)
    worksheet.set_column(29, 30, 15, format)
    worksheet.set_column(31, 32, 15, format_firma)

    worksheet.set_row(0, 75)

    # Escribir encabezado en hoja de cálculo
    for col_num, value in enumerate(df.columns.values):
        worksheet.write(0, col_num, value, header_format)

    # Escribir el resto de los datos en la hoja de cálculo
    for row_num, row_data in enumerate(df.values):
        for col_num, value in enumerate(row_data):
            worksheet.write(row_num + 1, col_num, None if value == 0.0000 else value)

    # Cerrar objeto ExcelWriter
    writer.close()

    # Crear respuesta HTTP
    output = io.BytesIO()
    with open('registros.xlsx', 'rb') as file:
        output.write(file.read())
    response = HttpResponse(
        output.getvalue(), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=registros.xlsx'
    return response
