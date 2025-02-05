from django.shortcuts import redirect, render
from perseovsfenix.logica_analisis import comparar_cantidades_concatenacion, obtener_diferencias_entre_items_pvf
import openpyxl
from django.http import JsonResponse
from io import BytesIO
from .models import Guia, NovedadPerseoVsFenix, NumeroActa, matfenix, matperseo
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages

def subir_pvf_matfenix(request):
    try:
        if request.method == 'POST' and 'file' in request.FILES:
            file = request.FILES['file']

            # Leer el archivo de Excel en memoria
            content = file.read()
            wb = openpyxl.load_workbook(
                filename=BytesIO(content), data_only=True)
            ws = wb.active  # Obtener la hoja activa

            row_count = 0

            # Empezamos desde la fila 2 para omitir el encabezado
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_count += 1

                # Obtener valores de las columnas según la posición indicada
                pedido = str(row[0])  # Columna A
                actividad = str(row[7])  # Columna H
                fecha = str(row[8])  # Columna I
                
                codigo = str(row[18]) if row[18] else str(
                    row[17])  # Columna S o R
                
                if codigo and codigo[-1] in ('A', 'P'):
                    codigo = codigo[:-1]
                # Columna S si no está vacía, de lo contrario, tomar columna R
                

                # Columna U, si está vacía, asignar cero
                # Columna U
                cantidad = row[20] if row[20] is not None else 0.00

                # Concatenación de pedido y código
                concatenacion = f"{pedido}-{codigo}"

                # Guardar en la base de datos
                matfenix.objects.create(
                    pedido=pedido,
                    actividad=actividad,
                    fecha=fecha,
                    codigo=codigo,
                    cantidad=cantidad,
                    concatenacion=concatenacion
                )
            messages.error(
                request,
                "Acta subida con exito."
            )
            return redirect('index_admin')
        # Si no es un POST, solo renderizamos la página
        return render(request, 'subir_pvf_matfenix.html')

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


def subir_pvf_matperseo(request):
    try:
        if request.method == 'POST' and 'file' in request.FILES:
            file = request.FILES['file']

            # Leer el archivo de Excel en memoria
            content = file.read()
            wb = openpyxl.load_workbook(
                filename=BytesIO(content), data_only=True)
            ws = wb.active  # Obtener la hoja activa

            row_count = 0

            # Empezamos desde la fila 2 para omitir el encabezado
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_count += 1

                # Obtener valores de las columnas según la posición indicada
                pedido = str(row[3])  # Columna A
                actividad = str(row[11])  # Columna H
                fecha = str(row[25])  # Columna I

                try:
                    
                    cod = Guia.objects.get(nombre_perseo=str(row[18]))
                    codigo=cod.nombre_fenix
                except ObjectDoesNotExist:
                    codigo = str(row[18])  # Si no existe, usa el valor de la columna

                # Columna U, si está vacía, asignar cero
                # Columna U
                cantidad = row[20] if row[20] is not None else 0.00

                # Columna U, si está vacía, asignar cero
                acta = row[26] if row[26] is not None else 0  # Columna U

                # Concatenación de pedido y código
                concatenacion = f"{pedido}-{codigo}"

                # Guardar en la base de datos
                matperseo.objects.create(
                    pedido=pedido,
                    actividad=actividad,
                    fecha=fecha,
                    codigo=codigo,
                    cantidad=cantidad,
                    concatenacion=concatenacion,
                    acta=acta
                )
                
            messages.error(
                request,
                "Extracción de Perseo subida con exito."
            )
            return redirect('index_admin')        
        # Si no es un POST, solo renderizamos la página
        return render(request, 'subir_pvf_matperseo.html')

    except Exception as e:
        print(e)
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

def reiniciar_bd_materiales(request):
    matfenix.objects.all().delete()
    matperseo.objects.all().delete()
    NovedadPerseoVsFenix.objects.all().delete()
    return render(request, "index.html")

def index(request):
    return render(request, "index.html")
 
def calculo_novedades_perseo_vs_fenix(request):
    obtener_diferencias_entre_items_pvf(request)
    comparar_cantidades_concatenacion()
    novedades = NovedadPerseoVsFenix.objects.all()
    return render(request, 'novedades_perseo_fenix.html', {'novedades': novedades})


def calculo_numero_acta():
    acta = NumeroActa.objects.first()
    print("acta")
    print(acta)
    pedidos_perseo = matperseo.objects.all()
    
    for pedido_perseo in pedidos_perseo:
        try:
            if str(pedido_perseo.acta) != str(acta.numero):
                faltante = NovedadPerseoVsFenix()
                faltante.concatenacion = pedido_perseo.concatenacion
                faltante.pedido = pedido_perseo.pedido
                faltante.actividad = pedido_perseo.actividad
                faltante.fecha = pedido_perseo.fecha
                faltante.codigo = pedido_perseo.codigo
                faltante.cantidad = pedido_perseo.cantidad
                faltante.observacion = "Acta incorrecta"
                faltante.acta = pedido_perseo.acta
                faltante.diferencia = 0
                faltante.save()
        except:
            print("error en el acta")

def reiniciar_novedades_perseo_vs_fenix(request):
    NovedadPerseoVsFenix.objects.all().delete()
    return render(request, 'novedades_perseo_fenix.html', {'novedades': ['']})

def novedades_perseo_vs_fenix(request):
    novedades = NovedadPerseoVsFenix.objects.all()
    return render(request, 'novedades_perseo_fenix.html', {'novedades': novedades})
