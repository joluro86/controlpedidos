from asyncio.windows_events import NULL
from email.policy import HTTP
from django.shortcuts import redirect, render
from django.urls import reverse
from django.contrib import messages
from django.db.models import Sum, Q, Count
from django.http import JsonResponse, HttpResponseRedirect
from openpyxl import load_workbook
from io import BytesIO
from gestionvencimientos.context_processors import numero_acta
from perseovsfenix.models import NumeroActa
from sellos.logica_informe import pedidos_de_instalado_a_series
from sellos.models import ActaSello, MaterialInstalado, SerieSello


def subir_acta(request, archivo):
    try:
        if request.method == 'POST':
            # Verificar si el archivo fue subido correctamente
            if 'file' not in request.FILES:
                return JsonResponse({'status': 'error', 'message': 'No file provided'}, status=400)
            file = request.FILES['file']
            archivo = request.POST.get('archivo')

            # Procesar el archivo en segundo plano (en este ejemplo, se hace sincrónicamente)
            process_extraccion(file, archivo)
            # Usamos HttpResponseRedirect en lugar de redirect
            return HttpResponseRedirect(reverse('home'))

        # Si no es un POST, solo renderizamos la página

        if archivo == "acta":
            extracc = "Acta"
        elif archivo == "series":
            extracc = "Series"
        elif archivo == "materiales":
            extracc = "Materiales instalados"
        else:
            messages.error(
                request,
                "Debe elegir la extracción a cargar."
            )
            return redirect('home')

        context = {
            'extraccion': extracc,
            'input_var': archivo
        }

        return render(request, 'subir_extracciones_sellos.html', context)

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


def process_extraccion(file, archivo):
    try:
        # Validar que el archivo es un Excel válido
        if not file.name.endswith('.xlsx'):
            raise ValueError("El archivo debe tener formato .xlsx")

        # Cargar el archivo en memoria usando BytesIO
        content = file.read()
        wb = load_workbook(filename=BytesIO(content))
        ws = wb.active  # Usamos la hoja activa por defecto
        row_count = 0

        # Comenzamos desde la fila 2
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_count += 1

            if archivo == "acta":
                # Crear una nueva instancia del modelo Acta
                registro = ActaSello()
                registro.pedido = row[0]
                registro.municipio = row[4]
                registro.actividad = row[7]
                registro.fecha = row[8]
                registro.pagina = row[9]
                registro.cantidad = row[20] if row[20] is not None else '0'

            if archivo == "series":
                # Crear una nueva instancia del modelo Acta
                registro = SerieSello()
                registro.pedido = "-"
                registro.consecutivo = row[16]
                registro.serie = row[3]
                registro.instalador = row[12]
                registro.va_en_informe = False

            if archivo == "materiales":
                # Crear una nueva instancia del modelo Acta
                registro = MaterialInstalado()
                registro.pedido = row[3]
                registro.consecutivo = row[2]
                registro.esta_en_acta = False

            try:
                registro.save()
            except Exception as e:
                print(f"Error al guardar la fila {row_count}: {e}")

    except Exception as e:
        print(f"Error al procesar el archivo: {e}")


def generar_informe(request):
    pedidos_de_instalado_a_series(request)
    return redirect('ver_informe_sellos')


def ver_informe_sellos(request):
    try:
        acta = NumeroActa.objects.filter().first()
    except Exception as e:
        print(e)
        acta = 0
        
    registros_sellos = SerieSello.objects.filter(va_en_informe=True)
    lista_sellos = []
    
    for sello in registros_sellos:
        lista_sellos.append({
            'actividad': ActaSello.objects.filter(pedido=sello.pedido).first().actividad,
            'serie': sello.serie,  # Suponiendo que existe este campo
            'pedido':sello.pedido,
            'municipio': ActaSello.objects.filter(pedido=sello.pedido).first().municipio,
            'fecha': ActaSello.objects.filter(pedido=sello.pedido).first().fecha,
            'pagina': ActaSello.objects.filter(pedido=sello.pedido).first().pagina,
            'instalador':sello.instalador,
            'acta':acta.numero
        })

    context = {
        'registros_sellos': lista_sellos,
    }

    return render(request, 'informe_sellos.html', context)

def novedades_sellos(request):
    pedidos_sellos = SerieSello.objects.filter(va_en_informe=True).values_list('pedido', flat=True)
    novedades_sellos= []
    
    for pedido in pedidos_sellos:  
                
        cantidad_series = SerieSello.objects.filter(pedido=pedido, va_en_informe=True).count()
        cantidad_acta = ActaSello.objects.filter(pedido=pedido).aggregate(total=Sum('cantidad'))['total'] or 0

        if (cantidad_series-cantidad_acta)!=0:

            novedades_sellos.append({
             'pedido':pedido,
             'cantidad_acta': cantidad_acta,
             'cantidad_series': cantidad_series,
             'novedad': 'Cantidad no concide'   
            }
            )
    
    pedidos_acta = ActaSello.objects.values_list('pedido', flat=True)

    for pedido in pedidos_acta:  
        
        if not any(novedad['pedido'] == pedido for novedad in novedades_sellos): 
                  
            cantidad_series = SerieSello.objects.filter(pedido=pedido, va_en_informe=True).count()
            cantidad_acta = ActaSello.objects.filter(pedido=pedido).aggregate(total=Sum('cantidad'))['total'] or 0

            if (cantidad_series-cantidad_acta)!=0:

                novedades_sellos.append({
                'pedido':pedido,
                'cantidad_acta': cantidad_acta,
                'cantidad_series': cantidad_series,
                'novedad': 'Cantidad no concide'   
                }
                )
        
    
    return render(request, 'novedades_sellos.html',{'novedades_sellos': novedades_sellos})

def reiniciar_extracciones(request):
    print("aqui")
    MaterialInstalado.objects.all().delete()
    ActaSello.objects.all().delete()
    SerieSello.objects.all().delete()
    return redirect('home')
